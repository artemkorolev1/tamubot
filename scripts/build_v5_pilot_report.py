#!/usr/bin/env python3
"""Build the v5 per-department pipeline report.

Same Excel layout as the v4/ISEN report — Summary + 4 findings sheets +
Stripped Headers — but reads from the v5 medallion tree at
``data/syllabi/<DEPT>/v5/``.

Sheets:
  - Summary           — 58 columns, ISEN-aligned
  - Content Pres.     — per-file v1 count + findings
  - Strip Compl.      — same shape
  - Structural        — same shape
  - Metadata          — same shape
  - Stripped Headers  — bullets of stripped boilerplate per file

Source artifacts (all in data/syllabi/<DEPT>/v5/):
  bronze/<stem>.md                       — hierarchy_depth, header_count
  silver/01_image_recovery/<stem>.md     — tokens in (BP input), image markers
  silver/02_false_positive/<stem>.md     — diff vs 01 → fp counts
  silver/03_boilerplate/<stem>.md        — tokens out
  silver/03_boilerplate/<stem>_stripped.txt — stripped header bullets
  silver/06_validate/<stem>_validation.json — full findings + counts
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_ROOT = Path("data/syllabi")

DARK_HDR_FILL = PatternFill(start_color="500000", end_color="500000", fill_type="solid")
BLUE_HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FONT_SMALL = Font(bold=True, color="FFFFFF", size=10)

CODE_TO_TERM = {"11": "Spring", "21": "Summer", "31": "Fall", "41": "Fall"}
IMAGE_MARKER = "<!-- image -->"
HEADER_RE = re.compile(r"^(#{1,6})\s+(.+)$", re.MULTILINE)


def _approx_tokens(text: str) -> int:
    return max(0, round(len(text) / 4))


def _parse_filename(stem: str) -> dict[str, str]:
    parts = stem.split("_")
    out = {"term": "", "course": "", "section": ""}
    if len(parts) >= 4 and parts[0].isdigit() and len(parts[0]) == 6:
        year = parts[0][:4]
        sem = CODE_TO_TERM.get(parts[0][4:], parts[0][4:])
        out["term"] = f"{sem} {year}"
        out["course"] = f"{parts[1]} {parts[2]}"
        out["section"] = parts[3]
    return out


def _count_headers(markdown: str) -> tuple[int, dict[int, int]]:
    depth: dict[int, int] = {}
    n = 0
    for m in HEADER_RE.finditer(markdown):
        lvl = len(m.group(1))
        depth[lvl] = depth.get(lvl, 0) + 1
        n += 1
    return n, depth


def _parse_stripped_sidecar(path: Path) -> list[dict]:
    """Parse the `[TYPE] header (level X, N chars)` bullets from the boilerplate
    sidecar produced by v4's filters/boilerplate.py."""
    if not path.exists():
        return []
    details = []
    pattern = re.compile(
        r"^\[(?P<type>[A-Z_]+)\]\s+(?P<header>.+?)\s+\(level\s+(?P<level>\d+),\s+(?P<chars>\d+)\s+chars\)"
    )
    for line in path.read_text(encoding="utf-8").splitlines():
        m = pattern.match(line)
        if m:
            details.append(
                {
                    "type": m.group("type"),
                    "header": m.group("header"),
                    "level": int(m.group("level")),
                    "chars": int(m.group("chars")),
                }
            )
    return details


def _count_headers_in_md(path: Path) -> set[str]:
    """Return the set of header texts in a markdown file (for FP diff)."""
    if not path.exists():
        return set()
    out = set()
    for m in HEADER_RE.finditer(path.read_text(encoding="utf-8")):
        out.add(m.group(2).strip())
    return out


def _load_manual_edits(dept: str) -> dict[str, list[dict]]:
    """Read manual_edits.csv if present. Returns {stem: [edit_dict, ...]}.

    Each edit_dict has keys: timestamp, stage, applied_by, llm_calls, summary.
    Multiple edits per stem stack in chronological order (CSV order).
    """
    path = DATA_ROOT / dept.upper() / "v5" / "logs" / "manual_edits.csv"
    edits: dict[str, list[dict]] = defaultdict(list)
    if not path.exists():
        return edits
    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            stem = (row.get("stem") or "").strip()
            if not stem:
                continue
            edits[stem].append(row)
    return edits


def collect(dept: str) -> tuple[dict[str, dict], dict[str, dict]]:
    v5 = DATA_ROOT / dept.upper() / "v5"
    bronze = v5 / "bronze"
    s01 = v5 / "silver" / "01_image_recovery"
    s02 = v5 / "silver" / "02_false_positive"
    s03 = v5 / "silver" / "03_boilerplate"
    s06 = v5 / "silver" / "06_validate"

    rows: dict[str, dict] = defaultdict(dict)
    manual_edits = _load_manual_edits(dept)

    def _empty_version() -> dict:
        return {
            "content_preservation": [],
            "strip_completeness": [],
            "structural_coherence": [],
            "metadata_accuracy": [],
            "proposals": {},
        }

    # validation[stem][version_label] -> per-category findings
    validation: dict[str, dict[str, dict]] = defaultdict(lambda: defaultdict(_empty_version))

    stems = sorted(p.stem for p in bronze.glob("*.md"))

    for stem in stems:
        # bronze: hierarchy_depth + token estimate (input to image_recovery)
        bronze_md = bronze / f"{stem}.md"
        if bronze_md.exists():
            txt = bronze_md.read_text(encoding="utf-8")
            n_h, depth = _count_headers(txt)
            rows[stem]["convert_headers"] = n_h
            rows[stem]["convert_hierarchy_depth"] = json.dumps(depth)
            rows[stem]["bronze_image_markers"] = txt.count(IMAGE_MARKER)

        # silver/01_image_recovery
        s01_md = s01 / f"{stem}.md"
        if s01_md.exists():
            txt = s01_md.read_text(encoding="utf-8")
            rows[stem]["s01_image_markers"] = txt.count(IMAGE_MARKER)
            rows[stem]["bp_tokens_in"] = _approx_tokens(txt)

        # image_recovery status: derive from marker drop
        before = rows[stem].get("bronze_image_markers", 0)
        after = rows[stem].get("s01_image_markers", 0)
        rows[stem]["img_before"] = before
        rows[stem]["img_after"] = after
        if before >= 2 and after < before:
            rows[stem]["img_status"] = "recovered"
        elif before < 2:
            rows[stem]["img_status"] = "skipped"
        else:
            rows[stem]["img_status"] = "passthrough"

        # silver/02_false_positive: derive FP demoted by diffing header sets vs s01
        s02_md = s02 / f"{stem}.md"
        if s01_md.exists() and s02_md.exists():
            h01 = _count_headers_in_md(s01_md)
            h02 = _count_headers_in_md(s02_md)
            removed = h01 - h02
            rows[stem]["fp_demoted"] = len(removed)
            len_diff = len(s01_md.read_text(encoding="utf-8")) - len(s02_md.read_text(encoding="utf-8"))
            rows[stem]["fp_cleanups"] = max(0, len_diff)

        # silver/03_boilerplate: tokens out + sidecar
        s03_md = s03 / f"{stem}.md"
        if s03_md.exists():
            rows[stem]["bp_tokens_out"] = _approx_tokens(s03_md.read_text(encoding="utf-8"))
        sidecar = s03 / f"{stem}_stripped.txt"
        details = _parse_stripped_sidecar(sidecar)
        rows[stem]["_bp_details"] = details
        rows[stem]["bp_sections_stripped"] = len(details)
        rows[stem]["bp_tokens_removed"] = sum(d["chars"] // 4 for d in details)

        # Hierarchy (always corrected in v5)
        rows[stem]["hier_status"] = "corrected"
        rows[stem]["hier_levels"] = rows[stem].get("convert_hierarchy_depth", "")

        # validation findings + counts — read versioned JSONs (_validation_v{N}.json)
        # Falls back to legacy unversioned file _validation.json as v1 if no _v1
        # file exists. Findings populate per-version columns; row aggregates use
        # the latest available version present on disk.
        version_files: list[tuple[str, Path]] = []
        for n in range(1, 9):
            p = s06 / f"{stem}_validation_v{n}.json"
            if p.exists():
                version_files.append((f"v{n}", p))
        if not version_files:
            legacy = s06 / f"{stem}_validation.json"
            if legacy.exists():
                version_files.append(("v1", legacy))

        for vlabel, vpath in version_files:
            try:
                data = json.loads(vpath.read_text(encoding="utf-8"))
            except Exception:
                continue
            findings = data.get("findings") or {}
            cats = ("content_preservation", "strip_completeness", "structural_coherence", "metadata_accuracy")
            for cat in cats:
                items = findings.get(cat) or []
                if items:
                    validation[stem][vlabel][cat] = items
                rows[stem][f"{vlabel}_val_{cat}"] = len(items)
            rows[stem][f"{vlabel}_val_total"] = sum(len(findings.get(c) or []) for c in cats)
            if data.get("proposals") and vlabel == version_files[-1][0]:
                # Keep proposals from the latest version only
                validation[stem]["proposals"] = data["proposals"]

        # Backwards-compat: val_* keys reflect latest version, for Summary cells
        # that aren't yet version-aware.
        if version_files:
            latest = version_files[-1][0]
            for cat in ("content_preservation", "strip_completeness", "structural_coherence", "metadata_accuracy"):
                rows[stem][f"val_{cat}"] = rows[stem].get(f"{latest}_val_{cat}", 0)
            rows[stem]["val_total"] = rows[stem].get(f"{latest}_val_total", 0)

        # Source detection: stems ending _HP are Howdy Portal
        rows[stem]["course_type"] = "Howdy Portal" if stem.endswith("_HP") else "Simple Syllabus"

        # Manual edits log (if any rows for this stem)
        if stem in manual_edits:
            rows[stem]["_manual_edits"] = manual_edits[stem]

    return rows, validation


# ── Sheet writers ─────────────────────────────────────────────────────────────


SUMMARY_HEADERS = [
    "File",
    "Course Type",
    "Term",
    "Course",
    "Section",
    "Tokens In",
    "Tokens Out",
    "Sections Stripped",
    "v1 Content",
    "v1 Strip",
    "v1 Structural",
    "v1 Metadata",
    "v1 Total",
    "v2 Content",
    "v2 Strip",
    "v2 Structural",
    "v2 Metadata",
    "v2 Total",
    "v3 Content",
    "v3 Strip",
    "v3 Structural",
    "v3 Metadata",
    "v3 Total",
    "v4 Content",
    "v4 Strip",
    "v4 Structural",
    "v4 Metadata",
    "v4 Total",
    "v5 Content",
    "v5 Strip",
    "v5 Structural",
    "v5 Metadata",
    "v5 Total",
    "Img Markers Before",
    "Img Markers After",
    "Gemini Recovered",
    "FP Headers Demoted",
    "FP Cleanups",
    "BP Sections Stripped",
    "BP Tokens In",
    "BP Tokens Out",
    "Hierarchy Corrected",
    "Heading Levels",
    "v6 Content",
    "v6 Strip",
    "v6 Structural",
    "v6 Metadata",
    "v6 Total",
    "v7 Content",
    "v7 Strip",
    "v7 Structural",
    "v7 Metadata",
    "v7 Total",
    "v8 Content",
    "v8 Strip",
    "v8 Structural",
    "v8 Metadata",
    "v8 Total",
    "Manual Edits",
]

SUMMARY_WIDTHS = {
    "File": 35,
    "Course Type": 16,
    "Term": 12,
    "Course": 10,
    "Section": 10,
    "Tokens In": 11,
    "Tokens Out": 12,
    "Sections Stripped": 14,
    "Img Markers Before": 14,
    "Img Markers After": 14,
    "Gemini Recovered": 14,
    "FP Headers Demoted": 14,
    "FP Cleanups": 11,
    "BP Sections Stripped": 14,
    "BP Tokens In": 11,
    "BP Tokens Out": 11,
    "Hierarchy Corrected": 14,
    "Heading Levels": 30,
    "Manual Edits": 60,
}

FINDINGS_HEADERS = [
    "File",
    "v1 Count",
    "v2 Count",
    "v3 Count",
    "v4 Count",
    "v5 Count",
    "v1 Findings",
    "v2 Findings",
    "v3 Findings",
    "v4 Findings",
    "v5 Findings",
    "v6 Count",
    "v6 Findings",
    "v7 Count",
    "v7 Findings",
    "v8 Count",
    "v8 Findings",
]

STRIPPED_HEADERS_COLS = ["File", "Sections Stripped", "Tokens In", "Tokens Out", "% Removed", "Stripped Headers"]


def _style_header_row(ws, ncols: int, fill: PatternFill, font: Font = HDR_FONT) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _write_summary(wb: Workbook, rows: dict[str, dict]) -> None:
    ws = wb.active
    ws.title = "Summary"
    for ci, label in enumerate(SUMMARY_HEADERS, start=1):
        ws.cell(1, ci, label)
    _style_header_row(ws, len(SUMMARY_HEADERS), DARK_HDR_FILL, HDR_FONT)

    def _ci(label: str) -> int:
        return SUMMARY_HEADERS.index(label) + 1

    ri = 2
    for stem in sorted(rows.keys()):
        data = rows[stem]
        fn = _parse_filename(stem)

        ws.cell(ri, _ci("File"), stem)
        ws.cell(ri, _ci("Course Type"), data.get("course_type", ""))
        ws.cell(ri, _ci("Term"), fn["term"])
        ws.cell(ri, _ci("Course"), fn["course"])
        ws.cell(ri, _ci("Section"), fn["section"])
        ws.cell(ri, _ci("Tokens In"), data.get("bp_tokens_in", ""))
        ws.cell(ri, _ci("Tokens Out"), data.get("bp_tokens_out", ""))
        ws.cell(ri, _ci("Sections Stripped"), data.get("bp_sections_stripped", ""))

        for n in range(1, 9):
            vlabel = f"v{n}"
            total_key = f"{vlabel}_val_total"
            if total_key not in data:
                continue
            ws.cell(ri, _ci(f"{vlabel} Content"), data.get(f"{vlabel}_val_content_preservation", ""))
            ws.cell(ri, _ci(f"{vlabel} Strip"), data.get(f"{vlabel}_val_strip_completeness", ""))
            ws.cell(ri, _ci(f"{vlabel} Structural"), data.get(f"{vlabel}_val_structural_coherence", ""))
            ws.cell(ri, _ci(f"{vlabel} Metadata"), data.get(f"{vlabel}_val_metadata_accuracy", ""))
            ws.cell(ri, _ci(f"{vlabel} Total"), data.get(total_key, ""))

        ws.cell(ri, _ci("Img Markers Before"), data.get("img_before", ""))
        ws.cell(ri, _ci("Img Markers After"), data.get("img_after", ""))
        status = data.get("img_status", "")
        gem_cell = "YES" if status == "recovered" else ("—" if status in ("skipped", "passthrough") else status)
        ws.cell(ri, _ci("Gemini Recovered"), gem_cell)
        if gem_cell == "YES":
            ws.cell(ri, _ci("Gemini Recovered")).fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )

        ws.cell(ri, _ci("FP Headers Demoted"), data.get("fp_demoted", ""))
        ws.cell(ri, _ci("FP Cleanups"), data.get("fp_cleanups", ""))
        ws.cell(ri, _ci("BP Sections Stripped"), data.get("bp_sections_stripped", ""))
        ws.cell(ri, _ci("BP Tokens In"), data.get("bp_tokens_in", ""))
        ws.cell(ri, _ci("BP Tokens Out"), data.get("bp_tokens_out", ""))

        ws.cell(ri, _ci("Hierarchy Corrected"), "YES" if data.get("hier_status") == "corrected" else "—")
        ws.cell(ri, _ci("Heading Levels"), str(data.get("hier_levels", "")))

        edits = data.get("_manual_edits") or []
        if edits:
            # Show each edit on its own line as: "[ts | by] summary"
            lines = [
                f"[{e.get('timestamp', '?')} | {e.get('applied_by', '?')}] {e.get('summary', '').strip()}"
                for e in edits
            ]
            cell = ws.cell(ri, _ci("Manual Edits"), "\n".join(lines))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # color per-version Total cells based on severity
        for n in range(1, 9):
            vlabel = f"v{n}"
            total = data.get(f"{vlabel}_val_total")
            if not isinstance(total, int):
                continue
            fill = None
            if total == 0:
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif total >= 8:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif total >= 3:
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            if fill:
                ws.cell(ri, _ci(f"{vlabel} Total")).fill = fill

        ri += 1

    for ci, label in enumerate(SUMMARY_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = SUMMARY_WIDTHS.get(label, 11)


def _write_findings_sheet(wb: Workbook, title: str, category: str, validation: dict[str, dict]) -> None:
    ws = wb.create_sheet(title=title)
    for ci, label in enumerate(FINDINGS_HEADERS, start=1):
        ws.cell(1, ci, label)
    _style_header_row(ws, len(FINDINGS_HEADERS), DARK_HDR_FILL, HDR_FONT)

    # Column index lookup. FINDINGS_HEADERS is: File, v1 Count..v5 Count, v1 Findings..v5 Findings, v6 Count, v6 Findings, v7 Count, v7 Findings, v8 Count, v8 Findings
    count_col = {f"v{n}": 1 + n for n in range(1, 6)}  # 2..6
    find_col = {f"v{n}": 6 + n for n in range(1, 6)}  # 7..11
    for n in range(6, 9):
        count_col[f"v{n}"] = 12 + (n - 6) * 2  # 12, 14, 16
        find_col[f"v{n}"] = 13 + (n - 6) * 2  # 13, 15, 17

    ri = 2
    for stem in sorted(validation.keys()):
        per_version = validation[stem]
        # per_version is a defaultdict that may also hold a "proposals" entry;
        # skip non-version keys.
        ws.cell(ri, 1, stem)
        max_findings = 0
        for vlabel, cats in per_version.items():
            if not vlabel.startswith("v"):
                continue
            findings = cats.get(category) or []
            ws.cell(ri, count_col[vlabel], len(findings))
            if findings:
                cell = ws.cell(ri, find_col[vlabel], "\n".join(f"• {f}" for f in findings))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                max_findings = max(max_findings, len(findings))
        if max_findings:
            ws.row_dimensions[ri].height = min(320, max(48, 16 * max_findings + 8))
        ri += 1

    ws.column_dimensions["A"].width = 38
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 10
    for c in range(7, 12):
        ws.column_dimensions[get_column_letter(c)].width = 60
    for c in range(12, 18):
        ws.column_dimensions[get_column_letter(c)].width = 12 if c % 2 == 0 else 60


def _write_stripped_headers(wb: Workbook, rows: dict[str, dict]) -> None:
    ws = wb.create_sheet(title="Stripped Headers")
    for ci, label in enumerate(STRIPPED_HEADERS_COLS, start=1):
        ws.cell(1, ci, label)
    _style_header_row(ws, len(STRIPPED_HEADERS_COLS), BLUE_HDR_FILL, HDR_FONT_SMALL)

    ri = 2
    for stem in sorted(rows.keys()):
        data = rows[stem]
        details = data.get("_bp_details") or []
        ws.cell(ri, 1, stem)
        ws.cell(ri, 2, data.get("bp_sections_stripped") or 0)
        ws.cell(ri, 3, data.get("bp_tokens_in") or 0)
        ws.cell(ri, 4, data.get("bp_tokens_out") or 0)
        ti = data.get("bp_tokens_in") or 0
        to = data.get("bp_tokens_out") or 0
        pct = round((1 - to / ti) * 100, 1) if ti else 0
        ws.cell(ri, 5, f"{pct}%")
        bullets = [f"• [{d['type']}] {d['header']}  ({d['chars']} chars)" for d in details]
        ws.cell(ri, 6, "\n".join(bullets) if bullets else "(none)")
        ws.cell(ri, 6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ri].height = min(280, max(40, 16 * len(details) + 8))
        ri += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 80


def build(dept: str, out_path: Path) -> tuple[int, int]:
    rows, validation = collect(dept)
    wb = Workbook()
    _write_summary(wb, rows)
    _write_findings_sheet(wb, "Content Pres.", "content_preservation", validation)
    _write_findings_sheet(wb, "Strip Compl.", "strip_completeness", validation)
    _write_findings_sheet(wb, "Structural", "structural_coherence", validation)
    _write_findings_sheet(wb, "Metadata", "metadata_accuracy", validation)
    _write_stripped_headers(wb, rows)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(rows), len(wb.sheetnames)


def main() -> None:
    p = argparse.ArgumentParser(description="Build v5 per-department pipeline report (ISEN-style layout)")
    p.add_argument("--dept", default="CSCE")
    p.add_argument("--output", default=None)
    args = p.parse_args()
    dept = args.dept.upper()
    out_path = Path(args.output) if args.output else DATA_ROOT / dept / "v5" / "silver" / "pipeline_v5_report.xlsx"
    n_rows, n_sheets = build(dept, out_path)
    print(f"Wrote {n_rows} {dept} rows across {n_sheets} sheets to {out_path}")


if __name__ == "__main__":
    main()
