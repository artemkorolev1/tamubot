#!/usr/bin/env python
"""Eval reuse helper — find unrun IDs and combine partial-batch reports.

Used by the run-eval skill to avoid re-running questions that already have
results. Two subcommands:

  missing GOLDEN_XLSX EXP_PREFIX
      Print a comma-separated list of question IDs that are present in the
      golden set but absent from every ``eval_<EXP_PREFIX>*.xlsx`` report in
      ``tamu_data/evals/reports/``. Empty output means everything is covered.

  combine EXP_PREFIX
      Glob ``eval_<EXP_PREFIX>*.xlsx`` reports, merge their Per-Query rows
      keyed by ``q#`` (later batches override earlier ones), and print a
      unified router-accuracy + RAGAS summary plus a per-question table.

Example — only run the gap, then summarize:

    GAP=$(python scripts/eval_delta.py missing \
        tamu_data/evals/golden_sets/ragas_20260519_curated20_v2.xlsx \
        router_subqueries_30cf2b2)
    make eval EXP=router_subqueries_30cf2b2_part2 IDS=$GAP ...
    python scripts/eval_delta.py combine router_subqueries_30cf2b2
"""

from __future__ import annotations

import argparse
import statistics
import sys
from pathlib import Path

import openpyxl

REPORTS_DIR = Path("tamu_data/evals/reports")


def _per_query_rows(xlsx: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    if "Per-Query" not in wb.sheetnames:
        return []
    ws = wb["Per-Query"]
    hdr = [c.value for c in ws[1]]
    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rd = dict(zip(hdr, row))
        if rd.get("q#") is not None:
            out.append(rd)
    return out


def _golden_ids(golden: Path) -> set:
    wb = openpyxl.load_workbook(golden, read_only=True)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    icol = hdr.index("id")
    return {row[icol] for row in ws.iter_rows(min_row=2, values_only=True) if row[icol] is not None}


def cmd_missing(golden: Path, exp_prefix: str) -> int:
    if not golden.exists():
        print(f"golden set not found: {golden}", file=sys.stderr)
        return 1
    done: set = set()
    files = sorted(REPORTS_DIR.glob(f"eval_{exp_prefix}*.xlsx"))
    for f in files:
        for r in _per_query_rows(f):
            done.add(r["q#"])
    total = _golden_ids(golden)
    missing = sorted(total - done, key=lambda x: (isinstance(x, str), x))
    extra = done - total
    if missing:
        print(",".join(str(x) for x in missing))
    print(
        f"  golden total={len(total)}  evaluated={len(done & total)}  missing={len(missing)}  "
        f"report files matching eval_{exp_prefix}*.xlsx: {len(files)}",
        file=sys.stderr,
    )
    if extra:
        print(f"  note: {len(extra)} IDs in reports but not in golden set: {sorted(extra)}", file=sys.stderr)
    return 0


def cmd_combine(exp_prefix: str) -> int:
    files = sorted(REPORTS_DIR.glob(f"eval_{exp_prefix}*.xlsx"))
    if not files:
        print(f"no report files matching eval_{exp_prefix}*.xlsx", file=sys.stderr)
        return 1
    rows_by_id: dict = {}
    for f in files:
        for r in _per_query_rows(f):
            rows_by_id[r["q#"]] = r  # later files override
    rows = sorted(rows_by_id.values(), key=lambda r: r["q#"])

    correct = sum(1 for r in rows if r.get("router_function_correct"))
    cps = [r["context_precision"] for r in rows if isinstance(r.get("context_precision"), (int, float))]
    crs = [r["context_recall"] for r in rows if isinstance(r.get("context_recall"), (int, float))]
    cited = sum(1 for r in rows if r.get("citation_pass"))
    cited_n = sum(1 for r in rows if r.get("citation_pass") is not None)

    print(f"merged {len(rows)} unique IDs from {len(files)} report file(s)")
    for f in files:
        print(f"  - {f.name}")
    print()
    print(f"  router accuracy:       {correct}/{len(rows)} = {100 * correct / len(rows):.1f}%")
    if cps:
        print(f"  mean context_precision: {statistics.mean(cps):.3f}  (n={len(cps)})")
    if crs:
        print(f"  mean context_recall:    {statistics.mean(crs):.3f}  (n={len(crs)})")
    if cited_n:
        print(f"  citation pass rate:     {cited}/{cited_n}")

    failures = [r for r in rows if not r.get("router_function_correct")]
    if failures:
        print()
        print(f"  router failures ({len(failures)}):")
        for r in failures:
            print(
                f"    Q{r['q#']:>3}  expected={r.get('expected_function')!s:>18}  got={r.get('router_function')!s:>18}"
            )

    print()
    print(f"  {'ID':>3} {'expected':>18} {'got':>18} {'ok':>3} {'chunks':>6} {'CP':>5} {'CR':>5}")
    print("  " + "-" * 70)
    for r in rows:
        cp = r.get("context_precision")
        cr = r.get("context_recall")
        cp_s = f"{cp:.2f}" if isinstance(cp, (int, float)) else "-"
        cr_s = f"{cr:.2f}" if isinstance(cr, (int, float)) else "-"
        mark = "✓" if r.get("router_function_correct") else "✗"
        print(
            f"  {r['q#']:>3} {r.get('expected_function', '')!s:>18} "
            f"{r.get('router_function', '')!s:>18} {mark:>3} "
            f"{r.get('chunks_retrieved', 0):>6} {cp_s:>5} {cr_s:>5}"
        )
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    m = sub.add_parser("missing", help="print IDs in the golden set with no row in any matching report")
    m.add_argument("golden", type=Path, help="path to golden-set xlsx")
    m.add_argument("exp_prefix", help="EXP prefix to match (matches eval_<prefix>*.xlsx)")

    c = sub.add_parser("combine", help="merge eval_<prefix>*.xlsx reports and print summary")
    c.add_argument("exp_prefix", help="EXP prefix to match")

    args = ap.parse_args()
    if args.cmd == "missing":
        return cmd_missing(args.golden, args.exp_prefix)
    return cmd_combine(args.exp_prefix)


if __name__ == "__main__":
    sys.exit(main())
