"""Full-STAT bake-off: run v6b (Docling+blocks) and v6c (ODL+repair) on every
STAT raw PDF and emit a side-by-side comparison at STAT/v6_full_compare.xlsx.

v6b bronze stems already on disk (blocks.json present) are skipped — Docling
output is deterministic, no need to redo. v6c is always re-run so every stem
goes through the post-rework letter-drop repair pass.

No LLM calls. Pure preprocessing comparison.
"""

from __future__ import annotations

import json
import re
import time
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from tamubot.ingestion.pipeline_v5.util import DATA_ROOT, hash_file
from tamubot.ingestion.pipeline_v6b import paths as v6b_paths
from tamubot.ingestion.pipeline_v6c import paths as v6c_paths
from tamubot.ingestion.pipeline_v6c.assets.bronze_odl import odl_convert

_HEADER_RE = re.compile(r"^#{1,6}\s+", re.MULTILINE)
_LETTER_DROP_TOKENS = (
    r"\bColege\b",
    r"\bSylabus\b",
    r"\bMeting\b",
    r"\bAditional\b",
    r"\bTextbok\b",
    r"\bwil\b",
    r"\bfal\b",
    r"\binstaling\b",
)
_LETTER_DROP_RE = re.compile("|".join(_LETTER_DROP_TOKENS))


def _count_md_headers(md_path: Path) -> int:
    if not md_path.exists():
        return 0
    return len(_HEADER_RE.findall(md_path.read_text(encoding="utf-8")))


def _run_v6b(stem: str, *, docling_converter=None) -> dict:
    """Process v6b for one stem.

    TEMP: when bronze artifacts are missing, spawns scripts/_v6b_one_stem.py
    as a subprocess so Docling/joblib memory is reclaimed by process exit.
    Without this, the 6 GB container OOM-kills after 7-17 stems.
    When Docker is bumped to 12+ GB, delete the subprocess branch and
    restore the in-process call.
    """
    import subprocess

    pdf = v6b_paths.raw_path(stem)
    bronze_dir = v6b_paths.bronze_blocks_path(stem).parent
    bronze_dir.mkdir(parents=True, exist_ok=True)

    blocks_path = v6b_paths.bronze_blocks_path(stem)
    md_path = v6b_paths.bronze_md_path(stem)

    if blocks_path.exists() and md_path.exists():
        blocks = json.loads(blocks_path.read_text(encoding="utf-8"))
        elapsed = None  # not regenerated
    else:
        t0 = time.time()
        result = subprocess.run(
            ["python", "-m", "scripts._v6b_one_stem", stem],
            capture_output=True,
            text=True,
            timeout=180,
        )
        elapsed = time.time() - t0
        if result.returncode != 0:
            raise RuntimeError(
                f"v6b subprocess failed for {stem}: rc={result.returncode}\n"
                f"stdout: {result.stdout[-500:]}\nstderr: {result.stderr[-500:]}"
            )
        blocks = json.loads(blocks_path.read_text(encoding="utf-8"))

    type_counts: Counter[str] = Counter(b["type"] for b in blocks)
    md_chars = len(md_path.read_text(encoding="utf-8")) if md_path.exists() else 0

    return {
        "v6b_blocks": len(blocks),
        "v6b_text_blocks": type_counts.get("text", 0),
        "v6b_heading_blocks": type_counts.get("heading", 0),
        "v6b_image_blocks": type_counts.get("image", 0),
        "v6b_table_blocks": type_counts.get("table", 0),
        "v6b_md_chars": md_chars,
        "v6b_md_headers": _count_md_headers(md_path),
        "v6b_md_sha": hash_file(md_path) if md_path.exists() else None,
        "v6b_time_s": round(elapsed, 2) if elapsed is not None else None,
    }


def _run_v6c(stem: str) -> dict:
    pdf = v6c_paths.raw_path(stem)
    bronze_dir = v6c_paths.bronze_md_path(stem).parent
    bronze_dir.mkdir(parents=True, exist_ok=True)

    result = odl_convert(pdf, output_dir=bronze_dir)

    md_out = v6c_paths.bronze_md_path(stem)
    md_out.write_text(result.markdown, encoding="utf-8")

    sidecar_out = v6c_paths.bronze_sidecar_path(stem)
    from tamubot.ingestion.pipeline_v5.schemas import HeadersSidecar

    sidecar = HeadersSidecar(headers=result.headers)
    sidecar_out.write_text(sidecar.model_dump_json(indent=2), encoding="utf-8")

    drops = len(_LETTER_DROP_RE.findall(result.markdown))
    return {
        "v6c_chars": len(result.markdown),
        "v6c_headers": len(result.headers),
        "v6c_md_headers": _count_md_headers(md_out),
        "v6c_pages": result.page_count,
        "v6c_time_s": round(result.timing_s, 2),
        "v6c_letter_drops": drops,
        "v6c_repair_word": result.repair_word_fixes,
        "v6c_repair_merge": result.repair_merge_fixes,
        "v6c_md_sha": hash_file(md_out),
    }


def _v5_chars(stem: str) -> int:
    p = DATA_ROOT / "STAT" / "v5" / "bronze" / f"{stem}.md"
    return len(p.read_text(encoding="utf-8")) if p.exists() else 0


def _stems() -> list[str]:
    raw_dir = DATA_ROOT / "STAT" / "v5" / "raw"
    return sorted(p.stem for p in raw_dir.glob("*.pdf"))


def _write_report(rows: list[dict], out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "v6b vs v6c (STAT 81)"

    headers = [
        "stem",
        "source",
        "pages",
        "v5_chars",
        "v6b_md_chars",
        "v6c_chars",
        "delta_v6c_v5_chars",
        "v6b_md_headers",
        "v6c_md_headers",
        "v6c_headers_struct",
        "v6b_blocks",
        "v6b_text",
        "v6b_heading",
        "v6b_image",
        "v6b_table",
        "v6c_letter_drops",
        "v6c_repair_word",
        "v6c_repair_merge",
        "v6b_time_s",
        "v6c_time_s",
        "v6b_md_sha",
        "v6c_md_sha",
        "md_sha_match_v5",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

    flag_drop = PatternFill("solid", fgColor="FFD9D9")
    flag_repair = PatternFill("solid", fgColor="FFF2CC")

    summary = {
        "n_total": 0,
        "n_hp": 0,
        "n_ss": 0,
        "v6b_time_total": 0.0,
        "v6c_time_total": 0.0,
        "v6b_chars_total": 0,
        "v6c_chars_total": 0,
        "v5_chars_total": 0,
        "letter_drops_total": 0,
        "repair_word_total": 0,
        "repair_merge_total": 0,
        "v6b_table_total": 0,
        "v6b_image_total": 0,
        "sha_match": 0,
    }

    for r in rows:
        v5_sha_p = DATA_ROOT / "STAT" / "v5" / "bronze" / f"{r['stem']}.md"
        v5_sha = hash_file(v5_sha_p) if v5_sha_p.exists() else None
        sha_match = (r.get("v6b_md_sha") == v5_sha) if v5_sha else False

        ws.append(
            [
                r["stem"],
                r["source"],
                r["v6c_pages"],
                r["v5_chars"],
                r["v6b_md_chars"],
                r["v6c_chars"],
                r["v6c_chars"] - r["v5_chars"],
                r["v6b_md_headers"],
                r["v6c_md_headers"],
                r["v6c_headers"],
                r["v6b_blocks"],
                r["v6b_text_blocks"],
                r["v6b_heading_blocks"],
                r["v6b_image_blocks"],
                r["v6b_table_blocks"],
                r["v6c_letter_drops"],
                r["v6c_repair_word"],
                r["v6c_repair_merge"],
                r["v6b_time_s"],
                r["v6c_time_s"],
                r["v6b_md_sha"],
                r["v6c_md_sha"],
                "YES" if sha_match else "NO",
            ]
        )
        if r["v6c_letter_drops"] > 0:
            for cell in ws[ws.max_row]:
                cell.fill = flag_drop
        elif r["v6c_repair_word"] + r["v6c_repair_merge"] > 0:
            for cell in ws[ws.max_row]:
                cell.fill = flag_repair

        summary["n_total"] += 1
        if r["source"] == "HowdyPortal":
            summary["n_hp"] += 1
        else:
            summary["n_ss"] += 1
        if r["v6b_time_s"]:
            summary["v6b_time_total"] += r["v6b_time_s"]
        summary["v6c_time_total"] += r["v6c_time_s"]
        summary["v6b_chars_total"] += r["v6b_md_chars"]
        summary["v6c_chars_total"] += r["v6c_chars"]
        summary["v5_chars_total"] += r["v5_chars"]
        summary["letter_drops_total"] += r["v6c_letter_drops"]
        summary["repair_word_total"] += r["v6c_repair_word"]
        summary["repair_merge_total"] += r["v6c_repair_merge"]
        summary["v6b_table_total"] += r["v6b_table_blocks"]
        summary["v6b_image_total"] += r["v6b_image_blocks"]
        if sha_match:
            summary["sha_match"] += 1

    for col_idx, name in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = max(12, min(28, len(name) + 2))
    ws.column_dimensions["A"].width = 40

    # Summary sheet
    ws2 = wb.create_sheet("summary")
    ws2.append(["metric", "value"])
    for k, v in summary.items():
        ws2.append([k, v])
    ws2.append([])
    ws2.append(["v6b_md == v5_md (%)", f"{100 * summary['sha_match'] / summary['n_total']:.1f}%"])
    ws2.append(
        ["v6c letter-drop rate (post-repair)", f"{summary['letter_drops_total']} hits across {summary['n_total']} docs"]
    )
    ws2.append(
        ["v6c repairs applied (word/merge)", f"{summary['repair_word_total']} / {summary['repair_merge_total']}"]
    )
    if summary["v6b_time_total"]:
        ws2.append(["v6b total wall-clock (new stems)", f"{summary['v6b_time_total']:.1f}s"])
    ws2.append(["v6c total wall-clock", f"{summary['v6c_time_total']:.1f}s"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 28

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def main() -> int:
    stems = _stems()
    print(f"STAT full bake-off: {len(stems)} stems (TEMP: v6b via subprocess, table-structure off)")
    rows: list[dict] = []
    t_run0 = time.time()
    for i, stem in enumerate(stems, 1):
        print(f"  [{i:>2}/{len(stems)}] {stem}", flush=True)
        v6b_metrics = _run_v6b(stem)
        v6c_metrics = _run_v6c(stem)
        rows.append(
            {
                "stem": stem,
                "source": "HowdyPortal" if stem.endswith("_HP") else "Simple Syllabus",
                "v5_chars": _v5_chars(stem),
                **v6b_metrics,
                **v6c_metrics,
            }
        )
    print(f"  total wall-clock: {time.time() - t_run0:.1f}s")

    out = DATA_ROOT / "STAT" / "v6_full_compare.xlsx"
    _write_report(rows, out)
    print(f"Report: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
