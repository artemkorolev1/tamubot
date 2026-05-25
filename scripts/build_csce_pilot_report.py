#!/usr/bin/env python3
"""Build a per-department pipeline report that mirrors the ISEN report's exact layout.

Sheets (matching data/syllabi/silver/docling_pipeline_v4_report.xlsx):
  - Summary           — 58 columns: file metadata, v1-v8 validation counts, per-step metrics
  - Content Pres.     — File, v1-v8 Count, v1-v8 Findings
  - Strip Compl.      — same shape
  - Structural        — same shape
  - Metadata          — same shape
  - Stripped Headers  — File, Sections Stripped, Tokens In, Tokens Out, % Removed, Stripped Headers

Pulls data from per-step CSV logs + per-file JSONs under both the dept-namespaced
(data/syllabi/<DEPT>/) and shared (data/syllabi/) trees.

For CSCE pilot we only have one effective version → v1 is populated, v2-v8 stay empty.
"""

from __future__ import annotations

import argparse
import ast
import csv
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_ROOT = Path("data/syllabi")
SHARED_LOGS = DATA_ROOT / "logs"
SHARED_SILVER = DATA_ROOT / "silver"

# Styling that matches the ISEN report
DARK_HDR_FILL = PatternFill(start_color="500000", end_color="500000", fill_type="solid")  # Summary + findings sheets
BLUE_HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Stripped Headers sheet
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FONT_SMALL = Font(bold=True, color="FFFFFF", size=10)


def _read_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _read_csv_layered(filename: str, dept: str) -> list[dict]:
    """Read CSV log from per-dept dir + shared dir.

    Returns entries ordered so the LATEST entry per file_stem comes last (caller
    overwrites earlier entries as it iterates). Deduplicates by 'file' keeping
    the row with the most recent 'timestamp', regardless of which dir it came from.
    """
    all_rows: list[dict] = []
    for base in (SHARED_LOGS, DATA_ROOT / dept.upper() / "logs"):
        all_rows.extend(_read_csv(base / filename))
    if not all_rows:
        return all_rows
    # Dedupe by 'file', keep most recent timestamp
    latest: dict[str, dict] = {}
    for row in all_rows:
        f = row.get("file")
        if not f:
            continue
        ts = row.get("timestamp") or ""
        if f not in latest or ts >= (latest[f].get("timestamp") or ""):
            latest[f] = row
    return list(latest.values())


def _silver_dirs(dept: str) -> list[Path]:
    return [DATA_ROOT / dept.upper() / "silver", SHARED_SILVER]


def _normalize_stem(name: str) -> str:
    for suf in (".md", ".pdf", ".json"):
        if name.endswith(suf):
            return name[: -len(suf)]
    return name


def _approx_tokens(text: str) -> int:
    return max(0, round(len(text) / 4))


CODE_TO_TERM = {"11": "Spring", "21": "Summer", "31": "Fall", "41": "Fall"}


def _parse_filename(stem: str) -> dict[str, str]:
    """Extract term/course/section from filename like 202611_CSCE_608_600_46648_v012."""
    parts = stem.split("_")
    out = {"term": "", "course": "", "section": ""}
    if len(parts) >= 4 and parts[0].isdigit() and len(parts[0]) == 6:
        year = parts[0][:4]
        sem_code = parts[0][4:]
        sem = CODE_TO_TERM.get(sem_code, sem_code)
        out["term"] = f"{sem} {year}"
        out["course"] = f"{parts[1]} {parts[2]}"
        out["section"] = parts[3]
    return out


def collect(dept: str) -> tuple[dict[str, dict], dict[str, dict]]:
    """Return (per_file, validation).

    per_file[stem]    -> dict of metrics (all the columns)
    validation[stem]  -> {category: [findings_strings], "proposals": {...}}
    """
    rows: dict[str, dict] = defaultdict(dict)
    validation: dict[str, dict] = defaultdict(lambda: {
        "content_preservation": [],
        "strip_completeness": [],
        "structural_coherence": [],
        "metadata_accuracy": [],
        "proposals": {},
    })
    dept_token = f"_{dept.upper()}_"

    for entry in _read_csv_layered("convert_log.csv", dept):
        stem = _normalize_stem(entry.get("file", ""))
        if dept_token not in stem:
            continue
        rows[stem]["convert_headers"] = entry.get("header_count")
        rows[stem]["convert_hierarchy_depth"] = entry.get("hierarchy_depth")

    for entry in _read_csv_layered("filter_image_recovery_log.csv", dept):
        stem = _normalize_stem(entry.get("file", ""))
        if dept_token not in stem:
            continue
        rows[stem]["img_before"] = entry.get("markers_before")
        rows[stem]["img_after"] = entry.get("markers_after")
        rows[stem]["img_status"] = entry.get("status")

    for entry in _read_csv_layered("filter_false_positive_log.csv", dept):
        stem = _normalize_stem(entry.get("file", ""))
        if dept_token not in stem:
            continue
        rows[stem]["fp_demoted"] = entry.get("demoted") or entry.get("demoted_count")
        rows[stem]["fp_cleanups"] = entry.get("cleanups_applied") or entry.get("text_cleanups_applied") or 0

    bp_details: dict[str, list[dict]] = {}
    for entry in _read_csv_layered("filter_boilerplate_log.csv", dept):
        stem = _normalize_stem(entry.get("file", ""))
        if dept_token not in stem:
            continue
        rows[stem]["bp_sections_stripped"] = entry.get("sections_stripped")
        rows[stem]["bp_tokens_removed"] = entry.get("tokens_removed")
        try:
            details = ast.literal_eval(entry.get("strip_details") or "[]")
            if isinstance(details, list):
                bp_details[stem] = details
        except (ValueError, SyntaxError):
            bp_details[stem] = []

    for entry in _read_csv_layered("filter_hierarchy_log.csv", dept):
        stem = _normalize_stem(entry.get("file", ""))
        if dept_token not in stem:
            continue
        rows[stem]["hier_status"] = entry.get("status")
        rows[stem]["hier_levels"] = entry.get("levels_after")

    for entry in _read_csv_layered("validate_log.csv", dept):
        stem = _normalize_stem(entry.get("file", ""))
        if dept_token not in stem:
            continue
        for cat in ("content_preservation", "strip_completeness", "structural_coherence", "metadata_accuracy"):
            if cat in entry and entry[cat]:
                rows[stem][f"val_{cat}"] = int(entry[cat]) if str(entry[cat]).isdigit() else entry[cat]
        if entry.get("total_issues"):
            rows[stem]["val_total"] = int(entry["total_issues"]) if str(entry["total_issues"]).isdigit() else entry["total_issues"]

    # Per-file validation JSONs — full findings
    for s in _silver_dirs(dept):
        for path in sorted((s / "05_validate").glob("*_validation.json")):
            stem = path.stem
            if stem.endswith("_validation"):
                stem = stem[: -len("_validation")]
            if dept_token not in stem:
                continue
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue
            findings = data.get("findings") or {}
            for cat in ("content_preservation", "strip_completeness", "structural_coherence", "metadata_accuracy"):
                if findings.get(cat):
                    validation[stem][cat] = findings[cat]
            if data.get("proposals"):
                validation[stem]["proposals"] = data["proposals"]

    # Enrichment
    for s in _silver_dirs(dept):
        for path in sorted((s / "05_enrich").glob("*.json")):
            stem = path.stem
            if dept_token not in stem:
                continue
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue
            cm = data.get("course_metadata") or {}
            rows[stem]["course_id"] = cm.get("course_id")
            rows[stem]["term_enr"] = cm.get("term")
            rows[stem]["section_enr"] = cm.get("section")
            rows[stem]["course_type"] = data.get("course_type")

    # Tokens in / tokens out: derive from silver/01_image_recovery (input to BP) and silver/03_boilerplate (output)
    for stem in list(rows.keys()):
        for base in _silver_dirs(dept):
            in_path = base / "01_image_recovery" / f"{stem}.md"
            out_path = base / "03_boilerplate" / f"{stem}.md"
            if in_path.exists():
                rows[stem].setdefault("bp_tokens_in", _approx_tokens(in_path.read_text(encoding="utf-8")))
            if out_path.exists():
                rows[stem].setdefault("bp_tokens_out", _approx_tokens(out_path.read_text(encoding="utf-8")))

    # Chunks
    for s in _silver_dirs(dept):
        for path in sorted((s / "06_chunk").glob("*.json")):
            stem = path.stem
            if dept_token not in stem:
                continue
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue
            chunks = data.get("chunks") or []
            rows[stem]["chunk_count"] = len(chunks)

    # Attach bp details
    for stem, dets in bp_details.items():
        rows[stem]["_bp_details"] = dets

    return rows, validation


# ── Sheet writers ─────────────────────────────────────────────────────────────


SUMMARY_HEADERS = [
    "File", "Course Type", "Term", "Course", "Section",
    "Tokens In", "Tokens Out", "Sections Stripped",
    "v1 Content", "v1 Strip", "v1 Structural", "v1 Metadata", "v1 Total",
    "v2 Content", "v2 Strip", "v2 Structural", "v2 Metadata", "v2 Total",
    "v3 Content", "v3 Strip", "v3 Structural", "v3 Metadata", "v3 Total",
    "v4 Content", "v4 Strip", "v4 Structural", "v4 Metadata", "v4 Total",
    "v5 Content", "v5 Strip", "v5 Structural", "v5 Metadata", "v5 Total",
    "Img Markers Before", "Img Markers After", "Gemini Recovered",
    "FP Headers Demoted", "FP Cleanups",
    "BP Sections Stripped", "BP Tokens In", "BP Tokens Out",
    "Hierarchy Corrected", "Heading Levels",
    "v6 Content", "v6 Strip", "v6 Structural", "v6 Metadata", "v6 Total",
    "v7 Content", "v7 Strip", "v7 Structural", "v7 Metadata", "v7 Total",
    "v8 Content", "v8 Strip", "v8 Structural", "v8 Metadata", "v8 Total",
]

SUMMARY_WIDTHS = {
    "File": 35, "Course Type": 14, "Term": 12, "Course": 10, "Section": 10,
    "Tokens In": 11, "Tokens Out": 12, "Sections Stripped": 14,
    "Img Markers Before": 14, "Img Markers After": 14, "Gemini Recovered": 14,
    "FP Headers Demoted": 14, "FP Cleanups": 11,
    "BP Sections Stripped": 14, "BP Tokens In": 11, "BP Tokens Out": 11,
    "Hierarchy Corrected": 14, "Heading Levels": 30,
}

FINDINGS_HEADERS = [
    "File",
    "v1 Count", "v2 Count", "v3 Count", "v4 Count", "v5 Count",
    "v1 Findings", "v2 Findings", "v3 Findings", "v4 Findings", "v5 Findings",
    "v6 Count", "v6 Findings",
    "v7 Count", "v7 Findings",
    "v8 Count", "v8 Findings",
]

STRIPPED_HEADERS_COLS = ["File", "Sections Stripped", "Tokens In", "Tokens Out", "% Removed", "Stripped Headers"]


def _style_header_row(ws, ncols: int, fill: PatternFill, font: Font = HDR_FONT) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _write_summary(wb: Workbook, rows: dict[str, dict]) -> None:
    ws = wb.active
    ws.title = "Summary"
    for ci, label in enumerate(SUMMARY_HEADERS, start=1):
        ws.cell(1, ci, label)
    _style_header_row(ws, len(SUMMARY_HEADERS), DARK_HDR_FILL, HDR_FONT)

    ri = 2
    for stem in sorted(rows.keys()):
        data = rows[stem]
        fn_parts = _parse_filename(stem)
        course = data.get("course_id") or fn_parts.get("course") or ""
        section = data.get("section_enr") or fn_parts.get("section") or ""
        term = data.get("term_enr") or fn_parts.get("term") or ""

        def _ci(label: str) -> int:
            return SUMMARY_HEADERS.index(label) + 1

        ws.cell(ri, _ci("File"), stem)
        ws.cell(ri, _ci("Course Type"), data.get("course_type") or "")
        ws.cell(ri, _ci("Term"), term)
        ws.cell(ri, _ci("Course"), course)
        ws.cell(ri, _ci("Section"), str(section))
        ws.cell(ri, _ci("Tokens In"), data.get("bp_tokens_in") or "")
        ws.cell(ri, _ci("Tokens Out"), data.get("bp_tokens_out") or "")
        ws.cell(ri, _ci("Sections Stripped"), data.get("bp_sections_stripped") or "")

        # v1 validation
        ws.cell(ri, _ci("v1 Content"), data.get("val_content_preservation") if data.get("val_content_preservation") is not None else "")
        ws.cell(ri, _ci("v1 Strip"), data.get("val_strip_completeness") if data.get("val_strip_completeness") is not None else "")
        ws.cell(ri, _ci("v1 Structural"), data.get("val_structural_coherence") if data.get("val_structural_coherence") is not None else "")
        ws.cell(ri, _ci("v1 Metadata"), data.get("val_metadata_accuracy") if data.get("val_metadata_accuracy") is not None else "")
        ws.cell(ri, _ci("v1 Total"), data.get("val_total") if data.get("val_total") is not None else "")
        # v2-v8 left empty (single effective run)

        # image_recovery
        ws.cell(ri, _ci("Img Markers Before"), data.get("img_before") or "")
        ws.cell(ri, _ci("Img Markers After"), data.get("img_after") or "")
        gem_status = data.get("img_status") or ""
        gem_cell = "YES" if gem_status == "recovered" else ("—" if gem_status in ("skipped", "pass_through") else gem_status or "")
        ws.cell(ri, _ci("Gemini Recovered"), gem_cell)
        if gem_cell == "YES":
            ws.cell(ri, _ci("Gemini Recovered")).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # FP
        ws.cell(ri, _ci("FP Headers Demoted"), data.get("fp_demoted") or "")
        ws.cell(ri, _ci("FP Cleanups"), data.get("fp_cleanups") or "")
        # BP detail
        ws.cell(ri, _ci("BP Sections Stripped"), data.get("bp_sections_stripped") or "")
        ws.cell(ri, _ci("BP Tokens In"), data.get("bp_tokens_in") or "")
        ws.cell(ri, _ci("BP Tokens Out"), data.get("bp_tokens_out") or "")

        # Hierarchy
        hier_status = data.get("hier_status") or ""
        ws.cell(ri, _ci("Hierarchy Corrected"), "YES" if hier_status == "corrected" else "—")
        ws.cell(ri, _ci("Heading Levels"), str(data.get("hier_levels") or ""))

        # Color the v1 Total cell
        total = data.get("val_total")
        if total is not None:
            try:
                t = int(total)
                fill = None
                if t == 0:
                    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif t >= 8:
                    fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif t >= 3:
                    fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                if fill:
                    ws.cell(ri, _ci("v1 Total")).fill = fill
            except (TypeError, ValueError):
                pass

        ri += 1

    # column widths
    for ci, label in enumerate(SUMMARY_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = SUMMARY_WIDTHS.get(label, 11)


def _write_findings_sheet(wb: Workbook, title: str, category: str, validation: dict[str, dict]) -> None:
    ws = wb.create_sheet(title=title)
    for ci, label in enumerate(FINDINGS_HEADERS, start=1):
        ws.cell(1, ci, label)
    _style_header_row(ws, len(FINDINGS_HEADERS), DARK_HDR_FILL, HDR_FONT)

    ri = 2
    for stem in sorted(validation.keys()):
        findings = validation[stem].get(category) or []
        ws.cell(ri, 1, stem)
        ws.cell(ri, 2, len(findings))  # v1 Count
        # v2..v8 counts stay blank
        # v1 Findings (col 7)
        if findings:
            ws.cell(ri, 7, "\n".join(f"• {f}" for f in findings))
            ws.cell(ri, 7).alignment = Alignment(wrap_text=True, vertical="top")
        ri += 1

    # widths
    ws.column_dimensions["A"].width = 38
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 10
    ws.column_dimensions["G"].width = 110  # v1 Findings
    for c in range(8, 12):
        ws.column_dimensions[get_column_letter(c)].width = 40
    for c in range(12, 18):
        ws.column_dimensions[get_column_letter(c)].width = 12 if c % 2 == 0 else 40

    # tall rows so wrapped findings are visible
    for r in range(2, ri):
        # height ~ 16 per finding, capped at 320
        n = (validation[ws.cell(r, 1).value] or {}).get(category) or []
        if n:
            ws.row_dimensions[r].height = min(320, max(48, 16 * len(n) + 8))


def _write_stripped_headers(wb: Workbook, rows: dict[str, dict]) -> None:
    ws = wb.create_sheet(title="Stripped Headers")
    for ci, label in enumerate(STRIPPED_HEADERS_COLS, start=1):
        ws.cell(1, ci, label)
    _style_header_row(ws, len(STRIPPED_HEADERS_COLS), BLUE_HDR_FILL, HDR_FONT_SMALL)

    ri = 2
    for stem in sorted(rows.keys()):
        data = rows[stem]
        details = data.get("_bp_details") or []
        ws.cell(ri, 1, stem)
        ws.cell(ri, 2, data.get("bp_sections_stripped") or 0)
        ws.cell(ri, 3, data.get("bp_tokens_in") or 0)
        ws.cell(ri, 4, data.get("bp_tokens_out") or 0)
        ti = data.get("bp_tokens_in") or 0
        to = data.get("bp_tokens_out") or 0
        pct = round((1 - to / ti) * 100, 1) if ti else 0
        ws.cell(ri, 5, f"{pct}%")
        bullets = []
        for d in details:
            header = d.get("header") if isinstance(d, dict) else str(d)
            btype = d.get("type") if isinstance(d, dict) else ""
            chars = d.get("chars") if isinstance(d, dict) else ""
            bullets.append(f"• [{btype}] {header}  ({chars} chars)")
        ws.cell(ri, 6, "\n".join(bullets) if bullets else "(none)")
        ws.cell(ri, 6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ri].height = min(280, max(40, 16 * len(details) + 8))
        ri += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 80


def build(dept: str, out_path: Path) -> tuple[int, int]:
    rows, validation = collect(dept)
    wb = Workbook()
    _write_summary(wb, rows)
    _write_findings_sheet(wb, "Content Pres.", "content_preservation", validation)
    _write_findings_sheet(wb, "Strip Compl.", "strip_completeness", validation)
    _write_findings_sheet(wb, "Structural", "structural_coherence", validation)
    _write_findings_sheet(wb, "Metadata", "metadata_accuracy", validation)
    _write_stripped_headers(wb, rows)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(rows), len(wb.sheetnames)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build per-department pipeline report (matches ISEN layout)")
    parser.add_argument("--dept", default="CSCE")
    parser.add_argument("--output", default=None)
    args = parser.parse_args()
    dept = args.dept.upper()
    out_path = (
        Path(args.output)
        if args.output
        else DATA_ROOT / dept / "silver" / "docling_pipeline_v4_report.xlsx"
    )
    n_rows, n_sheets = build(dept, out_path)
    print(f"Wrote {n_rows} {dept} rows across {n_sheets} sheets to {out_path}")


if __name__ == "__main__":
    main()
