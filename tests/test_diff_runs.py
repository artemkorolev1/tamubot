"""Tests for tamubot.evals.diff_runs — round-trip diff against a fixture xlsx."""

from pathlib import Path


def _make_golden_with_runs(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "id",
            "question",
            "reference_answer",
            "expected_function",
            "human_notes",
            "run:exp_a",
            "run:exp_b",
            "run:exp_a:ragas_faithfulness",
            "run:exp_b:ragas_faithfulness",
        ]
    )
    ws.append([1, "Q1?", "ref1", "hybrid_course", None, "ans_a_1", "ans_b_1", 0.6, 0.8])
    ws.append([2, "Q2?", "ref2", "semantic_general", None, "same", "same", 0.7, 0.7])
    ws.append([3, "Q3?", "ref3", "hybrid_course", None, "ans_a_3", "ans_b_3", 0.9, 0.4])
    wb.save(path)


def test_diff_runs_writes_xlsx_with_change_flags(tmp_path):
    from tamubot.evals.diff_runs import write_diff

    golden = tmp_path / "golden.xlsx"
    out = tmp_path / "diff.xlsx"
    _make_golden_with_runs(golden)

    write_diff(golden, "run:exp_a", "run:exp_b", out, metric=None)

    import openpyxl

    wb = openpyxl.load_workbook(out)
    assert "Diff" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    ws = wb["Diff"]
    headers = [c.value for c in ws[1]]
    assert headers[:5] == ["id", "question", "left_answer", "right_answer", "changed"]
    # row 2 = id 1, changed; row 3 = id 2, unchanged; row 4 = id 3, changed.
    # openpyxl serializes an empty string cell value as None on read-back.
    assert ws.cell(row=2, column=5).value == "Y"
    assert ws.cell(row=3, column=5).value in ("", None)
    assert ws.cell(row=4, column=5).value == "Y"


def test_diff_runs_with_metric_adds_delta(tmp_path):
    from tamubot.evals.diff_runs import write_diff

    golden = tmp_path / "golden.xlsx"
    out = tmp_path / "diff.xlsx"
    _make_golden_with_runs(golden)

    write_diff(golden, "run:exp_a", "run:exp_b", out, metric="ragas_faithfulness")

    import openpyxl

    wb = openpyxl.load_workbook(out)
    ws = wb["Diff"]
    headers = [c.value for c in ws[1]]
    assert "left_score" in headers
    assert "right_score" in headers
    assert "delta" in headers
    # id 1: 0.8 - 0.6 = 0.2 (improvement)
    delta_col = headers.index("delta") + 1
    assert abs(ws.cell(row=2, column=delta_col).value - 0.2) < 1e-6
    # id 3: 0.4 - 0.9 = -0.5 (regression)
    assert abs(ws.cell(row=4, column=delta_col).value + 0.5) < 1e-6


def test_diff_runs_missing_columns_exits(tmp_path):

    from tamubot.evals.diff_runs import write_diff

    golden = tmp_path / "golden.xlsx"
    out = tmp_path / "diff.xlsx"

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "question"])
    ws.append([1, "Q?"])
    wb.save(golden)

    try:
        write_diff(golden, "run:nope", "run:also_nope", out, metric=None)
    except SystemExit as e:
        assert e.code == 1
    else:
        raise AssertionError("Expected SystemExit when run columns missing")
