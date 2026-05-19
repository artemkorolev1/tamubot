"""RAGAS Testset Generator for TamuBot evaluation.

Generates document-grounded QA pairs from syllabus files using RAGAS's
TestsetGenerator and Knowledge Graph. Outputs an XLSX golden set compatible
with run_benchmark.py (same SCHEMA_COLUMNS).

Usage:
    python -m tamubot.evals.generate_ragas_testset --corpus-dir <path>
    python -m tamubot.evals.generate_ragas_testset --corpus-dir <path> --dry-run
    python -m tamubot.evals.generate_ragas_testset --corpus-dir <path> --testset-size 10
    python -m tamubot.evals.generate_ragas_testset --corpus-dir <path> --provider tamu
"""

from __future__ import annotations

import argparse
import difflib
import json
import logging
import re
from datetime import datetime
from pathlib import Path

from langchain_core.documents import Document

from tamubot.core import config

logger = logging.getLogger("tamubot.evals.ragas_testset")

# ---------------------------------------------------------------------------
# Synthesizer name → expected_function mapping
# ---------------------------------------------------------------------------

SYNTHESIZER_TO_FUNCTION: dict[str, str] = {
    "single_hop_specific_query_synthesizer": "hybrid_course",
    "multi_hop_specific_query_synthesizer": "recursive",
    "multi_hop_abstract_query_synthesizer": "semantic_general",
    "course_discovery_query_synthesizer": "semantic_general",
}


# Content-domain taxonomy distilled from `multi_persona_isen.yaml` personas.
# Used by CourseDiscoveryByTopic to generate "which course teaches X" questions.
# Keys are user-facing topic labels; values are keywords matched (case-insensitive,
# word-boundary) against CHUNK page_content to find covering courses.
TOPIC_TAXONOMY: dict[str, list[str]] = {
    "optimization": [
        "linear programming",
        "integer programming",
        "simplex method",
        "duality",
        "sensitivity analysis",
        "nonlinear programming",
        "dynamic programming",
        "network optimization",
        "Lagrange",
        "KKT",
        "operations research",
        "decision analysis",
        "constrained optimization",
        "unconstrained optimization",
        "Gurobi",
        "CPLEX",
        "mathematical programming",
    ],
    "human factors": [
        "human factors",
        "ergonomics",
        "biomechanics",
        "cognitive engineering",
        "sociotechnical",
        "musculoskeletal",
        "human performance",
        "human-systems integration",
        "man-machine",
        "work-system design",
        "human-centered",
        "cognitive workload",
        "situation awareness",
    ],
    "lean and manufacturing": [
        "lean manufacturing",
        "lean engineering",
        "lean thinking",
        "value stream mapping",
        "six sigma",
        "kaizen",
        "kanban",
        "CONWIP",
        "factory physics",
        "supply chain",
        "just-in-time",
        "DMAIC",
        "production system",
        "manufacturing system",
        "Toyota Production System",
    ],
    "data and statistics": [
        "design of experiments",
        "DOE",
        "ANOVA",
        "regression analysis",
        "hypothesis testing",
        "factorial design",
        "response surface",
        "applied statistics",
        "data analysis",
        "data science",
        "machine learning",
        "scikit-learn",
        "pandas",
        "statistical analysis",
        "model adequacy",
    ],
    "reliability and risk": [
        "reliability engineering",
        "reliability analysis",
        "risk analysis",
        "fault tree",
        "FMEA",
        "Markov model",
        "Weibull",
        "accelerated life testing",
        "probabilistic risk",
        "decision-making under uncertainty",
        "life-data analysis",
    ],
    "simulation": [
        "discrete event simulation",
        "Monte Carlo",
        "queueing theory",
        "stochastic process",
        "Markov chain",
        "Simio",
        "Arena",
        "AnyLogic",
        "simulation modeling",
        "variate generation",
        "random number generation",
    ],
    "materials and processes": [
        "mechanical behavior of materials",
        "finite element method",
        "FEM",
        "casting",
        "solidification",
        "metalworking",
        "deformation",
        "heat transfer",
        "manufacturing process",
    ],
    "systems engineering and management": [
        "systems engineering",
        "project management",
        "engineering management",
        "management of engineering",
        "leadership in engineering",
    ],
    "engineering economy": [
        "engineering economy",
        "capital investment",
        "depreciation",
        "discounted cash flow",
        "break-even analysis",
        "rate of return",
        "time value of money",
        "engineering economic analysis",
    ],
}


# Pre-compile per-topic regexes with word-boundaries to avoid false-positives
# (e.g. "LP" wouldn't match "help"; "DOE" matches as a word, not inside "DOES").
_TOPIC_REGEX: dict[str, "re.Pattern[str]"] = {
    topic: re.compile(
        r"\b(?:" + "|".join(re.escape(kw) for kw in kws) + r")\b",
        re.IGNORECASE,
    )
    for topic, kws in TOPIC_TAXONOMY.items()
}


COURSE_DISCOVERY_NUDGE = (
    "\n\nThe persona is choosing courses and wants to know which ones cover a specific "
    "topic. The 'combination' field names that topic. Generate a question that asks "
    "WHICH course(s) teach this topic (e.g., 'I want to study <topic>. Which course "
    "should I take?'), and an answer that ENUMERATES every course represented in the "
    "provided contexts by its course code (e.g., 'ISEN 620'), each with a one-sentence "
    "justification drawn from the syllabus content. Do not invent courses not in the "
    "contexts."
)

DISTRIBUTION_PRESETS = ("default", "balanced_50_50", "semantic_only", "course_coverage")

# Themes that describe administrative / boilerplate sections rather than
# course-content topics. Stripped from CHUNK nodes before generation so the
# multi-hop-abstract synthesizer can't pick them up as query subjects.
# Single-hop is unaffected (it iterates `entities`, not `themes`).
# Case-insensitive substring match.
BOILERPLATE_THEMES = (
    # Conduct / compliance
    "academic integrity",
    "academic misconduct",
    "academic honor",
    "honor code",
    "aggie code",
    "student conduct",
    "ada",
    "disability",
    "ferpa",
    "title ix",
    "mental health",
    "copyright",
    "ai use",
    "ai veracity",
    "ai technology",
    "ai-based inputs",
    "documentation of ai",
    # Grading / assessment admin
    "grading policy",
    "grading scale",
    "grading appeals",
    "grade breakdown",
    "grade determination",
    "re-grading",
    "regrading",
    "assignment weights",
    "assignment submission",
    "late work",
    "exam coverage",
    "comprehensive exam",
    # Schedule / participation admin
    "attendance",
    "excused absence",
    "participation",
    "course participation",
    "lecture activities",
    "lecture recordings",
    "discussion board",
    "communication guidelines",
    "weekly topics",
    "video modules",
    "course schedule",
    "course syllabus",
    "office hours",
    # Generic structural
    "course information",
    "instructor information",
    "teaching assistant",
    "instructor and teaching",
    "course materials",
    "course textbook",
    "course topics",
    "course grading components",
    "course learning outcomes",
    "optional instructional",
    "student responsibility",
    "reading responses",
    # Generic assessment nouns (still keep "Final Exam" because synthesizer benefits
    # from knowing assessment exists, but block pure "Exams"/"Homework"/"Assignments" wholesale themes
    # that don't carry domain content)
    "exams",
    "homework policy",
    "homework assignments",
    "homework",
    "assignments",
    "letter grades",
    "grade components",
    "grade assignment",
    "final project",
    "semester project",
    "semester-long",
    "project work",
    "research report",
    "distance learning",
    "class discussion",
    "deliverables",
    "documentation requirements",
)


def _install_theme_matching_logger() -> None:
    """Wrap ThemesPersonasMatchingPrompt.generate to log every call.

    Logs the themes passed in (per cluster), the personas, and the mapping
    returned. Empty mappings (the silent-zero failure mode) are flagged
    loudly. Diagnostic-only; do not leave on in production runs.
    """
    from ragas.testset.synthesizers.prompts import ThemesPersonasMatchingPrompt

    if getattr(ThemesPersonasMatchingPrompt, "_tamubot_wrapped", False):
        return

    original = ThemesPersonasMatchingPrompt.generate

    async def wrapped(self, data, llm, callbacks=None):
        result = await original(self, data=data, llm=llm, callbacks=callbacks)
        themes = list(data.themes) if hasattr(data, "themes") else []
        personas = [p.name for p in data.personas] if hasattr(data, "personas") else []
        mapping = getattr(result, "mapping", {}) or {}
        empty_personas = [p for p in personas if not mapping.get(p)]
        logger.info(
            "[theme-match] themes=%s personas=%s mapping=%s",
            themes[:20],
            personas,
            {k: v for k, v in mapping.items()},
        )
        if empty_personas:
            logger.warning(
                "[theme-match] EMPTY mapping for personas=%s; themes were %s",
                empty_personas,
                themes,
            )
        return result

    ThemesPersonasMatchingPrompt.generate = wrapped  # type: ignore[assignment]
    ThemesPersonasMatchingPrompt._tamubot_wrapped = True  # type: ignore[attr-defined]
    logger.info("[theme-match] logger installed on ThemesPersonasMatchingPrompt")


def filter_boilerplate_themes(kg) -> None:
    """Strip boilerplate themes from CHUNK nodes in-place.

    Mutates `node.properties["themes"]` to drop entries that match any
    BOILERPLATE_THEMES substring (case-insensitive). Logs counts.
    """
    blocked = [b.lower() for b in BOILERPLATE_THEMES]
    total_before = 0
    total_after = 0
    chunks_emptied = 0
    for node in kg.nodes:
        if node.type.name != "CHUNK":
            continue
        themes = node.properties.get("themes")
        if not themes:
            continue
        total_before += len(themes)
        kept = [t for t in themes if not any(b in t.lower() for b in blocked)]
        total_after += len(kept)
        if kept != themes:
            node.properties["themes"] = kept
            if not kept:
                chunks_emptied += 1
    dropped = total_before - total_after
    logger.info(
        "Boilerplate filter: dropped %d theme entries (%d → %d); %d chunks left themeless",
        dropped,
        total_before,
        total_after,
        chunks_emptied,
    )


SELECTION_TIME_NUDGE = (
    "\n\nThe student is choosing courses for a future semester and is not "
    "currently enrolled. Prefer questions about durable course attributes — "
    "topics covered, learning outcomes, tools used, grading structure, "
    "prerequisites. Avoid questions about specific term-bound deadlines "
    "(e.g., 'when is the midterm scheduled' or 'what is due next week')."
)


# Date patterns that indicate a transient (term-bound) reference.
# Items whose 'user_input' question matches any of these are dropped.
_TRANSIENT_DATE_PATTERNS = [
    # "on Tuesday, February 24" / "on February 24" / "Friday, May 1"
    re.compile(
        r"\b(on|by)\s+(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)?,?\s*(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2}",
        re.IGNORECASE,
    ),
    # Standalone month + day with year suffix ("February 24, 2026")
    re.compile(
        r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+(20\d{2})",
        re.IGNORECASE,
    ),
    # Relative deadlines
    re.compile(
        r"\b(next|this|last|coming)\s+(week|month|monday|tuesday|wednesday|thursday|friday|saturday|sunday)\b",
        re.IGNORECASE,
    ),
    # Compact dates
    re.compile(r"\b\d{1,2}/\d{1,2}/(?:\d{2}|\d{4})\b"),
    # Year-month-day
    re.compile(r"\b\d{4}-\d{1,2}-\d{1,2}\b"),
    # Day-letter + M/D with no year ("T 10/6", "R 11/10", "F 9/3")
    re.compile(r"\b[MTWRFSU]\s+\d{1,2}/\d{1,2}\b"),
    # Lecture/Lec. + number ("Lec. 12", "Lecture 12")
    re.compile(r"\bLec(?:ture|\.)?\s+\d+\b", re.IGNORECASE),
    # Homework deadline phrases ("Homework #3 due")
    re.compile(r"\bHomework\s+#?\d+\s+(?:is\s+)?due\b", re.IGNORECASE),
    # Exam scheduling phrases ("Exam 3 is scheduled", "Exam 3 scheduled")
    re.compile(r"\bExam\s+\d+\s+(?:is\s+)?scheduled\b", re.IGNORECASE),
]


# Phrases that frame two contexts as a single-course internal contradiction.
# Used to reject multi-CRN multi-hop items where the LLM has invented a
# "discrepancy" by conflating two different courses' policies.
_SINGLE_COURSE_FRAMING_PATTERNS = [
    re.compile(r"\bdiscrepancy\b", re.IGNORECASE),
    re.compile(r"\binconsistency\b", re.IGNORECASE),
    re.compile(r"\b(one|another|the\s+other|a\s+different)\s+section\b", re.IGNORECASE),
    re.compile(r"\bmentioned\s+elsewhere\b", re.IGNORECASE),
    re.compile(r"\blisted\s+in\s+one\b", re.IGNORECASE),
    re.compile(r"\bis\s+the\s+correct\b", re.IGNORECASE),
    re.compile(r"\bconflicting\s+\w+", re.IGNORECASE),
]


# Course-code injection: detect already-named course codes, and detect stems
# that need a course identifier prepended.
_COURSE_CODE_RE = re.compile(r"\b[A-Z]{2,5}\s*\d{3,4}[A-Z]?(?:/\d{3,4})?\b")
_NEEDS_COURSE_CONTEXT_RE = re.compile(
    r"\b(this|the)\s+(course|class|syllabus|assignments?|final\s+exam|midterm|readings?|"
    r"grading(?:\s+policy)?|policy|exam)\b",
    re.IGNORECASE,
)


def _has_transient_date(text: str) -> bool:
    """Return True if `text` mentions a specific date or relative deadline."""
    if not isinstance(text, str) or not text:
        return False
    return any(p.search(text) for p in _TRANSIENT_DATE_PATTERNS)


def _has_single_course_framing(text: str) -> bool:
    """Return True if `text` frames its content as a single-course contradiction.

    Used to reject multi-CRN items whose stem asserts internal consistency
    (e.g. "in one section vs. elsewhere", "the discrepancy") — those are
    almost always synthesizer hallucinations conflating two courses.
    """
    if not isinstance(text, str) or not text:
        return False
    return any(p.search(text) for p in _SINGLE_COURSE_FRAMING_PATTERNS)


def _apply_selection_time_nudge(synth) -> None:
    """Append SELECTION_TIME_NUDGE to the synthesizer's query-generation prompt."""
    prompts = synth.get_prompts()
    prompt = prompts["query_answer_generation_prompt"]
    prompt.instruction = prompt.instruction + SELECTION_TIME_NUDGE
    synth.set_prompts(**{"query_answer_generation_prompt": prompt})


def _apply_course_discovery_nudge(synth) -> None:
    """Append COURSE_DISCOVERY_NUDGE to the synthesizer's prompt.

    Tries common Ragas prompt names; logs which one was patched. Safe no-op
    if no matching prompt is found (the synth will fall back to default
    generation behavior, which usually still produces sensible output given
    the topic in `combination`).
    """
    try:
        prompts = synth.get_prompts()
    except Exception as e:  # noqa: BLE001
        logger.warning("course-discovery nudge: get_prompts failed: %s", e)
        return

    for key in ("query_answer_generation_prompt", "abstract_query_prompt"):
        if key in prompts:
            prompt = prompts[key]
            prompt.instruction = prompt.instruction + COURSE_DISCOVERY_NUDGE
            synth.set_prompts(**{key: prompt})
            logger.info("course-discovery nudge applied to prompt '%s'", key)
            return
    logger.warning(
        "course-discovery nudge: no recognized prompt key on synth (available: %s)",
        list(prompts.keys()),
    )


# ---------------------------------------------------------------------------
# Step 1 — Document loading
# ---------------------------------------------------------------------------

# Filename pattern: {TERM}_{DEPT}_{COURSENUM}_{SECTION}_{CRN}_v{VERSION}.json
_FILENAME_RE = re.compile(
    r"^(?P<term>\d+)_(?P<dept>[A-Z]+)_(?P<coursenum>\d+)_(?P<section>\d+)_(?P<crn>\d+)_v\d+\.json$"
)


def load_corpus(corpus_dir: Path, include_terms: set[str] | None = None) -> list[Document]:
    """Load v3_step3_flat / v4 06_chunk JSON files as LangChain Documents.

    Each document gets the full concatenated Markdown as page_content.
    RAGAS handles its own internal segmentation during KG construction.

    If `include_terms` is provided, only files whose filename starts with one
    of those term codes (e.g. {"202611", "202641"}) are loaded.
    """
    json_files = sorted(corpus_dir.glob("*.json"))
    if not json_files:
        raise FileNotFoundError(f"No JSON files found in {corpus_dir}")

    docs: list[Document] = []
    for fp in json_files:
        if fp.name.startswith("_"):
            continue  # skip metadata files like _run_meta_*.json
        if include_terms is not None:
            term_prefix = fp.name.split("_", 1)[0]
            if term_prefix not in include_terms:
                continue

        with open(fp) as f:
            data = json.load(f)

        chunks = data.get("chunks", [])
        if not chunks:
            logger.warning("Skipping %s — no chunks", fp.name)
            continue

        page_content = "\n\n".join(c["content"] for c in chunks)

        meta = data.get("course_metadata", {})
        crn = meta.get("crn")
        course_id = meta.get("course_id", "")
        term = meta.get("term", "")
        section = meta.get("section", "")

        # Fallback: parse CRN from filename if metadata has null CRN
        if not crn:
            m = _FILENAME_RE.match(fp.name)
            if m:
                crn = m.group("crn")
                if not course_id:
                    course_id = f"{m.group('dept')} {m.group('coursenum')}"

        docs.append(
            Document(
                page_content=page_content,
                metadata={
                    "crn": crn or "",
                    "course_id": course_id,
                    "term": term,
                    "section": section,
                    "source_file": fp.name,
                },
            )
        )

    logger.info("Loaded %d documents from %s", len(docs), corpus_dir)
    return docs


# ---------------------------------------------------------------------------
# Step 2 — LLM setup
# ---------------------------------------------------------------------------


def build_llm(provider: str, temperature: float):
    """Create a RAGAS-compatible LLM via llm_factory.

    Ragas 0.4.x requires a client instance — text-only mode was removed.
    Google uses the Instructor adapter wrapping google.genai.Client.
    """
    from ragas.llms import llm_factory

    if provider == "google":
        from google import genai

        client = genai.Client(api_key=config.GOOGLE_API_KEY)
        return llm_factory(
            "gemini-2.0-flash",
            provider="google",
            client=client,
            temperature=temperature,
        )

    if provider == "tamu":
        from openai import OpenAI

        from tamubot.evals.tamu_openai_workaround import wrap_for_tamu

        client = wrap_for_tamu(
            OpenAI(
                api_key=config.TAMU_API_KEY,
                base_url=config.TAMU_BASE_URL,
            )
        )
        return llm_factory(
            config.TAMU_MODEL,
            provider="openai",
            client=client,
            temperature=temperature,
        )

    raise ValueError(f"Unknown provider: {provider!r}. Use 'google' or 'tamu'.")


# ---------------------------------------------------------------------------
# Step 3 — Embedding setup
# ---------------------------------------------------------------------------


def build_embeddings():
    """Google gemini-embedding-001 wrapped for RAGAS.

    Note: text-embedding-004 was removed from Google's v1beta endpoint
    in 2026. gemini-embedding-001 is the current replacement (3072-dim).
    TAMU's gateway has no embedding endpoint, so embeddings always go
    to Google direct regardless of --provider.
    """
    from langchain_google_genai import GoogleGenerativeAIEmbeddings
    from ragas.embeddings import LangchainEmbeddingsWrapper

    return LangchainEmbeddingsWrapper(
        GoogleGenerativeAIEmbeddings(
            model="models/gemini-embedding-001",
            google_api_key=config.GOOGLE_API_KEY,
        )
    )


# ---------------------------------------------------------------------------
# Step 4 — Knowledge Graph construction
# ---------------------------------------------------------------------------


def _build_syllabus_ner_prompt():
    """NER prompt tuned for syllabus entities, ignoring shared boilerplate."""
    from ragas.testset.transforms.extractors.llm_based import (
        NEROutput,
        NERPrompt,
        TextWithExtractionLimit,
    )

    class SyllabusNERPrompt(NERPrompt):
        instruction: str = (
            "Extract named entities from the given syllabus text. "
            "Focus on: course IDs (e.g. CSCE 638), instructor names, "
            "grading components and weights, prerequisites, textbooks, "
            "exam formats, project types, tools/software, credit hours, "
            "and specific deadlines or dates.\n\n"
            "IGNORE the following boilerplate that is identical across syllabi: "
            "honor code statements, ADA/disability policy, university attendance policy, "
            "FERPA notice, mental health resources, Title IX statements, "
            "and academic integrity definitions. "
            "Do not extract entities from these sections.\n\n"
            "Limit output to the top entities. "
            "Ensure the number of entities does not exceed the specified maximum."
        )
        examples: list[tuple[TextWithExtractionLimit, NEROutput]] = [
            (
                TextWithExtractionLimit(
                    text=(
                        "CSCE 638 - Section 600 - Algorithms\n"
                        "Instructor: Dr. Fang Song\n"
                        "Prerequisites: CSCE 411\n"
                        "Grading: Homework 40%, Midterm 25%, Final 35%\n"
                        "Textbook: Introduction to Algorithms by Cormen et al.\n"
                        "An Aggie does not lie, cheat, or steal."
                    ),
                    max_num=10,
                ),
                NEROutput(
                    entities=[
                        "CSCE 638",
                        "Dr. Fang Song",
                        "CSCE 411",
                        "Homework 40%",
                        "Midterm 25%",
                        "Final 35%",
                        "Introduction to Algorithms",
                        "Cormen et al.",
                    ]
                ),
            ),
        ]

    return SyllabusNERPrompt()


def build_transforms(llm, embedding_model, documents: list[Document]):
    """Custom KG transforms adapted from RAGAS default_transforms.

    Uses a syllabus-specific NER prompt to avoid boilerplate cross-links.
    """
    from ragas.testset.graph import NodeType
    from ragas.testset.transforms import Parallel
    from ragas.testset.transforms.extractors import (
        EmbeddingExtractor,
        HeadlinesExtractor,
        NERExtractor,
        SummaryExtractor,
    )
    from ragas.testset.transforms.extractors.llm_based import ThemesExtractor
    from ragas.testset.transforms.filters import CustomNodeFilter
    from ragas.testset.transforms.relationship_builders import (
        CosineSimilarityBuilder,
        OverlapScoreBuilder,
    )
    from ragas.testset.transforms.splitters import HeadlineSplitter
    from ragas.utils import num_tokens_from_string

    def filter_long_docs(node):
        return node.type == NodeType.DOCUMENT and num_tokens_from_string(node.properties.get("page_content", "")) > 500

    def filter_docs(node):
        return node.type == NodeType.DOCUMENT

    def filter_chunks(node):
        return node.type == NodeType.CHUNK

    # Check document length distribution to decide transform pipeline
    token_counts = [num_tokens_from_string(doc.page_content) for doc in documents]
    long_pct = sum(1 for t in token_counts if t > 500) / len(token_counts)

    if long_pct < 0.25:
        raise ValueError(
            f"Only {long_pct:.0%} of documents exceed 500 tokens. "
            "Syllabus documents should be longer — check your corpus."
        )

    headline_extractor = HeadlinesExtractor(llm=llm, filter_nodes=lambda node: filter_long_docs(node))
    # Splitter must skip short docs that HeadlinesExtractor never set
    # 'headlines' on — otherwise it raises ValueError and crashes the run.
    splitter = HeadlineSplitter(min_tokens=500, filter_nodes=lambda node: filter_long_docs(node))
    summary_extractor = SummaryExtractor(llm=llm, filter_nodes=lambda node: filter_long_docs(node))
    node_filter = CustomNodeFilter(llm=llm, filter_nodes=lambda node: filter_chunks(node))

    summary_emb_extractor = EmbeddingExtractor(
        embedding_model=embedding_model,
        property_name="summary_embedding",
        embed_property_name="summary",
        filter_nodes=lambda node: filter_long_docs(node),
    )
    theme_extractor = ThemesExtractor(llm=llm, filter_nodes=lambda node: filter_chunks(node))
    ner_extractor = NERExtractor(
        llm=llm,
        filter_nodes=lambda node: filter_chunks(node),
        prompt=_build_syllabus_ner_prompt(),
    )

    cosine_sim_builder = CosineSimilarityBuilder(
        property_name="summary_embedding",
        new_property_name="summary_similarity",
        threshold=0.7,
        filter_nodes=lambda node: filter_long_docs(node),
    )
    ner_overlap_sim = OverlapScoreBuilder(threshold=0.01, filter_nodes=lambda node: filter_chunks(node))

    return [
        headline_extractor,
        splitter,
        summary_extractor,
        node_filter,
        Parallel(summary_emb_extractor, theme_extractor, ner_extractor),
        Parallel(cosine_sim_builder, ner_overlap_sim),
    ]


def build_or_load_kg(
    documents: list[Document],
    llm,
    embedding_model,
    kg_path: Path,
    rebuild: bool,
):
    """Load cached KG or build from scratch."""
    from ragas.testset.graph import KnowledgeGraph, Node, NodeType
    from ragas.testset.transforms import apply_transforms

    if kg_path.exists() and not rebuild:
        logger.info("Loading cached KG from %s", kg_path)
        return KnowledgeGraph.load(str(kg_path))

    logger.info("Building Knowledge Graph from %d documents...", len(documents))
    kg = KnowledgeGraph()
    for doc in documents:
        kg.nodes.append(
            Node(
                type=NodeType.DOCUMENT,
                properties={
                    "page_content": doc.page_content,
                    "document_metadata": doc.metadata,
                },
            )
        )
    logger.info("Seeded KG with %d DOCUMENT nodes", len(kg.nodes))
    transforms = build_transforms(llm, embedding_model, documents)
    apply_transforms(kg, transforms)
    kg_path.parent.mkdir(parents=True, exist_ok=True)
    kg.save(str(kg_path))
    logger.info("KG saved to %s", kg_path)
    return kg


# ---------------------------------------------------------------------------
# Step 5 — Query distribution
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Custom synthesizers — bypass Ragas's diversification caps
# ---------------------------------------------------------------------------


def _build_custom_synthesizers(llm):
    """Build single-hop + multi-hop synthesizers that maximize source-doc coverage.

    Default Ragas synthesizers break at n samples after iterating chunks in
    insertion order; with 122 chunks and n=10, only the first ~10 chunks ever
    get sampled. These custom variants iterate one chunk per source document
    (single-hop) and every entities_overlap pair as a triplet (multi-hop).
    """
    import random
    import typing as t
    from dataclasses import dataclass, field

    import numpy as np
    from ragas.testset.synthesizers.multi_hop.base import (
        MultiHopQuerySynthesizer,
    )
    from ragas.testset.synthesizers.prompts import (
        ThemesPersonasMatchingPrompt,
    )
    from ragas.testset.synthesizers.single_hop.specific import (
        SingleHopSpecificQuerySynthesizer,
    )

    @dataclass
    class CourseCoverageSingleHop(SingleHopSpecificQuerySynthesizer):
        """Iterates one chunk per source_file, picking the entity-richest chunk.

        Guarantees N distinct source documents in N samples (up to the number
        of docs in the KG). Bypasses the default break-at-n-after-insertion-order
        behavior that caused the same first ~10 chunks to dominate every run.
        """

        name: str = "single_hop_specific_query_synthesizer"

        async def _generate_scenarios(self, n, knowledge_graph, persona_list, callbacks):
            child_rels = [r for r in knowledge_graph.relationships if r.type == "child"]
            chunk_to_doc = {r.target.id: r.source for r in child_rels}

            # Group chunks by source_file (best chunk per file = most entities)
            per_source_best: dict[str, t.Any] = {}
            for node in knowledge_graph.nodes:
                if node.type.name != "CHUNK":
                    continue
                ents = self._extract_themes_from_items(node.properties.get(self.property_name, []))
                if not ents:
                    continue
                parent = chunk_to_doc.get(node.id)
                if not parent:
                    continue
                src = parent.properties.get("document_metadata", {}).get("source_file", "")
                if not src:
                    continue
                cur = per_source_best.get(src)
                if cur is None or len(ents) > len(
                    self._extract_themes_from_items(cur.properties.get(self.property_name, []))
                ):
                    per_source_best[src] = node

            ordered_chunks = list(per_source_best.values())
            random.shuffle(ordered_chunks)
            logger.info("CourseCoverageSingleHop: %d distinct-source chunks available", len(ordered_chunks))

            scenarios: list = []
            samples_per_node = max(1, int(np.ceil(n / max(len(ordered_chunks), 1))))
            for node in ordered_chunks:
                if len(scenarios) >= n:
                    break
                themes = self._extract_themes_from_items(node.properties.get(self.property_name, []))
                if not themes:
                    continue
                # Bypass theme_persona_matching (expensive + fails at max_tokens with many personas).
                # Associate every persona with every theme on this chunk.
                forced_mapping = {p.name: themes for p in persona_list}
                base_scenarios = self.prepare_combinations(
                    node, themes, personas=persona_list, persona_concepts=forced_mapping
                )
                scenarios.extend(self.sample_combinations(base_scenarios, samples_per_node))
            return scenarios

    @dataclass
    class CourseCoveragePairwiseMultiHop(MultiHopQuerySynthesizer):
        """Iterates every entities_overlap pair as a triplet, not aggregated clusters.

        find_n_indirect_clusters caps at 10 hub-anchored templates with our KG.
        find_two_nodes_single_rel returns 572 distinct pairs — far more course
        diversity. Each pair becomes a 2-hop scenario.
        """

        name: str = "multi_hop_abstract_query_synthesizer"
        theme_persona_matching_prompt: ThemesPersonasMatchingPrompt = field(
            default_factory=ThemesPersonasMatchingPrompt
        )

        async def _generate_scenarios(self, n, knowledge_graph, persona_list, callbacks):
            from ragas.testset.synthesizers.base import QueryLength, QueryStyle

            results = list(
                knowledge_graph.find_two_nodes_single_rel(
                    relationship_condition=lambda rel: rel.type == "entities_overlap"
                )
            )
            if not results:
                logger.warning("CourseCoveragePairwiseMultiHop: no entities_overlap pairs found")
                return []

            # Map chunks to their parent doc CRN so we can filter to cross-CRN pairs
            child_rels = [r for r in knowledge_graph.relationships if r.type == "child"]
            chunk_to_crn: dict = {}
            for r in child_rels:
                crn = r.source.properties.get("document_metadata", {}).get("crn", "")
                chunk_to_crn[r.target.id] = crn

            blocked = [b.lower() for b in BOILERPLATE_THEMES]

            def _useful_themes(triplet):
                _, rel, _ = triplet
                items = rel.properties.get("overlapped_items", []) or []
                kept = []
                for pair in items:
                    for token in pair:
                        if isinstance(token, str) and not any(b in token.lower() for b in blocked):
                            kept.append(token)
                return list(dict.fromkeys(kept))

            cross_crn_triplets = []
            same_crn_dropped = 0
            for triplet in results:
                node_a, _, node_b = triplet
                crn_a = chunk_to_crn.get(node_a.id, "")
                crn_b = chunk_to_crn.get(node_b.id, "")
                # Require cross-CRN pair (different courses)
                if crn_a == crn_b:
                    same_crn_dropped += 1
                    continue
                themes = _useful_themes(triplet)
                if themes:
                    cross_crn_triplets.append((triplet, themes))

            random.shuffle(cross_crn_triplets)
            logger.info(
                "CourseCoveragePairwiseMultiHop: %d cross-CRN pairs (dropped %d same-CRN, %d total)",
                len(cross_crn_triplets),
                same_crn_dropped,
                len(results),
            )

            num_per_triplet = max(1, n // max(len(cross_crn_triplets), 1))
            scenarios: list = []
            styles = list(QueryStyle)
            lengths = list(QueryLength)
            for triplet, themes in cross_crn_triplets:
                if len(scenarios) >= n:
                    break
                node_a, _, node_b = triplet
                # Construct MultiHopScenarios DIRECTLY — bypass prepare_combinations
                # because its valid_nodes filter drops node_b when the overlap term
                # isn't a verbatim entity match (e.g. "Homework (25%)" vs "Homework 25%").
                candidate_samples = []
                for combination in [[th] for th in themes]:
                    for persona in persona_list:
                        for style in styles:
                            for length in lengths:
                                candidate_samples.append(
                                    {
                                        "combination": tuple(combination),
                                        "persona": persona,
                                        "nodes": [node_a, node_b],
                                        "style": style,
                                        "length": length,
                                    }
                                )
                random.shuffle(candidate_samples)
                selected = candidate_samples[:num_per_triplet]
                scenarios.extend(self.convert_to_scenario(s) for s in selected)
            return scenarios

    @dataclass
    class CourseDiscoveryByTopic(MultiHopQuerySynthesizer):
        """Topic → courses discovery scenarios ("Which course teaches X?").

        For each topic in TOPIC_TAXONOMY, scans every CHUNK's page_content for
        keyword matches (case-insensitive, word-boundary). Groups matches by
        parent DOCUMENT, keeping the best-matching chunk per doc. Topics
        covered by ≥2 documents become scenarios; each scenario contains one
        anchor chunk per matching course. The LLM is prompted (via
        _apply_course_discovery_nudge) to enumerate every covering course in
        the answer.
        """

        name: str = "course_discovery_query_synthesizer"
        theme_persona_matching_prompt: ThemesPersonasMatchingPrompt = field(
            default_factory=ThemesPersonasMatchingPrompt
        )

        async def _generate_scenarios(self, n, knowledge_graph, persona_list, callbacks):
            from ragas.testset.synthesizers.base import QueryLength, QueryStyle

            child_rels = [r for r in knowledge_graph.relationships if r.type == "child"]
            chunk_to_doc = {r.target.id: r.source for r in child_rels}

            topic_scenarios: list[tuple[str, list]] = []
            for topic, regex in _TOPIC_REGEX.items():
                # Group by course_id (not source_file) so multiple sections of
                # the same course collapse to one anchor — discovery questions
                # need distinct courses, not distinct CRNs.
                course_best: dict[str, tuple] = {}
                for node in knowledge_graph.nodes:
                    if node.type.name != "CHUNK":
                        continue
                    content = node.properties.get("page_content", "")
                    matches = len({m.group(0).lower() for m in regex.finditer(content)})
                    if matches == 0:
                        continue
                    parent = chunk_to_doc.get(node.id)
                    if not parent:
                        continue
                    course_id = parent.properties.get("document_metadata", {}).get("course_id", "")
                    if not course_id:
                        continue
                    cur = course_best.get(course_id)
                    if cur is None or matches > cur[1]:
                        course_best[course_id] = (node, matches)
                if len(course_best) >= 2:
                    anchor_chunks = [info[0] for info in course_best.values()]
                    topic_scenarios.append((topic, anchor_chunks))

            logger.info(
                "CourseDiscoveryByTopic: %d topics with ≥2 matching courses: %s",
                len(topic_scenarios),
                {t: len(chunks) for t, chunks in topic_scenarios},
            )

            random.shuffle(topic_scenarios)
            styles = list(QueryStyle)
            lengths = list(QueryLength)
            scenarios: list = []
            for topic, anchor_chunks in topic_scenarios:
                if len(scenarios) >= n:
                    break
                for persona in persona_list:
                    if len(scenarios) >= n:
                        break
                    scenarios.append(
                        self.convert_to_scenario(
                            {
                                "combination": (topic,),
                                "persona": persona,
                                "nodes": anchor_chunks,
                                "style": random.choice(styles),
                                "length": random.choice(lengths),
                            }
                        )
                    )
            return scenarios

    return (
        CourseCoverageSingleHop(llm=llm),
        CourseCoveragePairwiseMultiHop(llm=llm),
        CourseDiscoveryByTopic(llm=llm),
    )


def build_query_distribution(llm, preset: str = "default", multi_hop_relation: str = "entities_overlap_score"):
    """Return a Ragas query_distribution list for the chosen preset.

    Every SingleHopSpecificQuerySynthesizer in the returned list has
    SELECTION_TIME_NUDGE appended to its query_answer_generation_prompt
    instruction.

    `multi_hop_relation` controls which KG edge property the multi-hop-abstract
    synthesizer walks to find clusters. ``entities_overlap_score`` (default)
    produces ~10 distinct topical clusters; ``summary_similarity`` saturates to
    2 hub-anchored clusters because syllabi share too much structural language.

    Presets:
      - "default":         50% single-hop / 30% multi-hop-specific / 20% multi-hop-abstract
      - "balanced_50_50":  50% single-hop-specific / 50% multi-hop-abstract
      - "semantic_only":   100% multi-hop-abstract (all -> expected_function=semantic_general)
    """
    from ragas.testset.synthesizers.multi_hop.abstract import (
        MultiHopAbstractQuerySynthesizer,
    )
    from ragas.testset.synthesizers.multi_hop.specific import (
        MultiHopSpecificQuerySynthesizer,
    )
    from ragas.testset.synthesizers.single_hop.specific import (
        SingleHopSpecificQuerySynthesizer,
    )

    def _single_hop():
        s = SingleHopSpecificQuerySynthesizer(llm=llm)
        _apply_selection_time_nudge(s)
        return s

    def _multi_hop_abstract():
        return MultiHopAbstractQuerySynthesizer(llm=llm, relation_property=multi_hop_relation)

    if preset == "balanced_50_50":
        return [
            (_single_hop(), 0.50),
            (_multi_hop_abstract(), 0.50),
        ]
    if preset == "semantic_only":
        return [(_multi_hop_abstract(), 1.0)]
    if preset == "course_coverage":
        # Custom synthesizers: single-hop iterates 1-per-doc, pairwise iterates
        # entities_overlap pairs cross-CRN, discovery iterates content-topic
        # taxonomy mapping topics → covering courses.
        cc_single, cc_multi, cc_discovery = _build_custom_synthesizers(llm)
        _apply_selection_time_nudge(cc_single)
        _apply_course_discovery_nudge(cc_discovery)
        return [
            (cc_single, 0.40),
            (cc_multi, 0.40),
            (cc_discovery, 0.20),
        ]
    # default
    return [
        (_single_hop(), 0.50),
        (MultiHopSpecificQuerySynthesizer(llm=llm), 0.30),
        (_multi_hop_abstract(), 0.20),
    ]


# ---------------------------------------------------------------------------
# Step 6 — Generation
# ---------------------------------------------------------------------------


def generate_testset(kg, llm, embedding_model, query_distribution, testset_size: int, persona_list=None):
    """Run RAGAS TestsetGenerator and return a DataFrame.

    If ``persona_list`` is a non-empty list of ragas.testset.persona.Persona
    objects, it is passed to the TestsetGenerator constructor (Ragas 0.4.x API);
    otherwise Ragas falls back to its internal default persona generation.
    """
    from ragas.testset import TestsetGenerator

    generator = TestsetGenerator(
        llm=llm,
        embedding_model=embedding_model,
        knowledge_graph=kg,
        persona_list=persona_list if persona_list else None,
    )
    testset = generator.generate(
        testset_size=testset_size,
        query_distribution=query_distribution,
        raise_exceptions=False,
    )
    return testset.to_pandas()  # type: ignore[union-attr]


# ---------------------------------------------------------------------------
# Step 7 — Validation
# ---------------------------------------------------------------------------


def validate_testset(df, documents: list[Document], min_ratio: float = 0.3, kg=None):
    """Drop rows whose reference_contexts cannot be traced to source docs.

    Uses fuzzy string matching against the original document content.
    If ``kg`` is provided, also drops cross-CRN items whose stem uses
    single-course framing language (false-premise multi-hops).
    Returns the filtered DataFrame.
    """
    source_texts = {doc.metadata["source_file"]: doc.page_content for doc in documents}
    all_content = "\n\n".join(source_texts.values())

    chunk_index = _build_chunk_crn_index(kg) if kg is not None else None

    keep_mask = []
    for idx, row in df.iterrows():
        contexts = row.get("reference_contexts")
        if not contexts or (isinstance(contexts, list) and len(contexts) == 0):
            logger.debug("Row %s: empty reference_contexts — dropping", idx)
            keep_mask.append(False)
            continue

        if isinstance(contexts, str):
            try:
                contexts = json.loads(contexts)
            except json.JSONDecodeError:
                contexts = [contexts]

        synthesizer = (row.get("synthesizer_name") or "").lower()
        is_multi_hop_abstract = "multi_hop_abstract" in synthesizer
        is_course_discovery = "course_discovery" in synthesizer

        if is_multi_hop_abstract or is_course_discovery:
            # Multi-hop-abstract / discovery reference_contexts may be
            # paraphrased or aggregated across multiple chunks; trust Ragas's
            # grounding rather than fuzzy-matching against source text.
            matched = bool(contexts)  # only drop if contexts is empty
        else:
            # Existing single-hop / multi-hop-specific path: fuzzy match against
            # source corpus.
            matched = False
            for ctx in contexts:
                if not isinstance(ctx, str) or not ctx.strip():
                    continue
                ratio = difflib.SequenceMatcher(None, ctx[:500], all_content).quick_ratio()
                if ratio >= min_ratio:
                    matched = True
                    break
                # Substring fallback: if any 50-char window of the context
                # appears verbatim in the source corpus, accept it.
                ctx_clean = ctx.strip()
                for start in range(0, max(1, len(ctx_clean) - 50), 25):
                    snippet = ctx_clean[start : start + 50]
                    if len(snippet) >= 50 and snippet in all_content:
                        matched = True
                        break
                if matched:
                    break

        # Drop items whose question references a specific calendar date.
        question = row.get("user_input") or ""
        if matched and _has_transient_date(question):
            logger.debug("Row %s: question contains a transient date — dropping", idx)
            matched = False

        # Also drop if the reference answer itself contains transient-date
        # markers (e.g. "T 10/6", "Lec. 12") — the expected response would
        # be tied to a specific term. Don't scan contexts: syllabi routinely
        # include schedule sections, so context dates would over-filter.
        if matched:
            answer = row.get("reference") or ""
            if _has_transient_date(str(answer)):
                logger.debug(
                    "Row %s: reference answer contains a transient date — dropping",
                    idx,
                )
                matched = False

        # Cross-CRN single-course-framing check: reject items that pair contexts
        # from different courses but frame them as one course's internal
        # contradiction ("discrepancy", "one section vs elsewhere").
        if matched and chunk_index is not None and not is_course_discovery:
            crns = _attribute_crns(contexts, chunk_index)
            if len(crns) > 1 and _has_single_course_framing(question):
                logger.debug(
                    "Row %s: cross-CRN item with single-course framing — dropping",
                    idx,
                )
                matched = False

        keep_mask.append(matched)

    before = len(df)
    df_clean = df[keep_mask].reset_index(drop=True)
    dropped = before - len(df_clean)
    logger.info("Validation: kept %d / %d items (dropped %d)", len(df_clean), before, dropped)
    return df_clean


# ---------------------------------------------------------------------------
# Step 7b — Near-duplicate dedup
# ---------------------------------------------------------------------------


def _find_chunk_for_context(ref_context: str, kg) -> object | None:
    """Locate the CHUNK node whose page_content contains this reference_context.

    Strips the optional ``<N-hop>`` prefix, then probes the first 100 chars.
    Returns the matched Node or None.
    """
    if not isinstance(ref_context, str):
        return None
    ctx = _HOP_PREFIX_RE.sub("", ref_context).strip()
    if not ctx:
        return None
    probe = ctx[:100]
    for node in kg.nodes:
        if node.type.name != "CHUNK":
            continue
        if probe in node.properties.get("page_content", ""):
            return node
    return None


def mask_used_nodes(kg, seed_df, mask_whole_crn: bool = False) -> int:
    """Empty themes + entities on CHUNK nodes already used by `seed_df`.

    With ``mask_whole_crn=False`` (default), only chunks that directly
    produced a seed row are masked. With ``mask_whole_crn=True``, every
    chunk belonging to a doc whose CRN appears in the seed is masked — much
    more aggressive, pushes Ragas into uncovered courses entirely.

    Returns the number of chunks blanked.
    """
    # Build chunk → CRN lookup once
    child_rels = [r for r in kg.relationships if r.type == "child"]
    chunk_to_doc = {r.target.id: r.source for r in child_rels}

    def _chunk_crn(node):
        parent = chunk_to_doc.get(node.id)
        if not parent:
            return ""
        return parent.properties.get("document_metadata", {}).get("crn", "") or ""

    used_chunks: set[str] = set()
    used_crns: set[str] = set()
    for _, row in seed_df.iterrows():
        rc = row.get("reference_contexts")
        if isinstance(rc, str):
            try:
                rc = json.loads(rc)
            except json.JSONDecodeError:
                rc = [rc]
        if not isinstance(rc, list):
            continue
        for ctx in rc:
            node = _find_chunk_for_context(ctx, kg)
            if node is not None:
                used_chunks.add(node.id)
                crn = _chunk_crn(node)
                if crn:
                    used_crns.add(crn)

    masked = 0
    for node in kg.nodes:
        if node.type.name != "CHUNK":
            continue
        if mask_whole_crn:
            if _chunk_crn(node) in used_crns:
                node.properties["themes"] = []
                node.properties["entities"] = []
                masked += 1
        else:
            if node.id in used_chunks:
                node.properties["themes"] = []
                node.properties["entities"] = []
                masked += 1

    logger.info(
        "Node masking (whole_crn=%s): blanked %d CHUNK nodes; %d CRNs covered by seed",
        mask_whole_crn,
        masked,
        len(used_crns),
    )
    return masked


def cap_per_crn(df, kg, max_per_crn: int = 3):
    """Drop rows beyond `max_per_crn` items sharing the same source CRN.

    Ragas re-draws scenarios from the same chunks across batches, so a single
    information-rich course (e.g., ISEN 608 with its full textbook + exam
    schedule) can dominate single-hop output. This cap forces coverage across
    courses by dropping the later-generated rows once the per-CRN budget is
    exhausted.
    """
    if "reference_contexts" not in df.columns or len(df) == 0:
        return df

    chunk_index = _build_chunk_crn_index(kg)
    seen: dict[str, int] = {}
    keep_mask: list[bool] = []
    for _, row in df.iterrows():
        rc = row.get("reference_contexts")
        if isinstance(rc, str):
            try:
                rc = json.loads(rc)
            except json.JSONDecodeError:
                rc = [rc]
        crn = _attribute_crn(rc, chunk_index)
        if not crn:
            keep_mask.append(True)
            continue
        seen[crn] = seen.get(crn, 0) + 1
        keep_mask.append(seen[crn] <= max_per_crn)

    before = len(df)
    df_clean = df[keep_mask].reset_index(drop=True)
    dropped = before - len(df_clean)
    logger.info("Per-CRN cap (max=%d): kept %d / %d (dropped %d)", max_per_crn, len(df_clean), before, dropped)
    return df_clean


def deduplicate_by_question_similarity(df, embedding_model, threshold: float = 0.85):
    """Drop rows whose `user_input` is a near-duplicate of an earlier row.

    Embeds each question via the same Ragas-wrapped embedder used for the KG.
    For any pair (i, j) with i < j and cosine(q_i, q_j) >= threshold, drops j.
    Earlier rows are preferred so chronological/batch order is preserved.
    """
    if len(df) <= 1:
        return df

    questions = [str(q or "") for q in df["user_input"].tolist()]
    raw = embedding_model.embed_documents(questions)
    import numpy as np

    vecs = np.asarray(raw, dtype=float)
    norms = np.linalg.norm(vecs, axis=1, keepdims=True)
    norms[norms == 0] = 1.0
    vecs = vecs / norms

    sim = vecs @ vecs.T
    n = len(questions)
    keep_mask = [True] * n
    for j in range(n):
        if not keep_mask[j]:
            continue
        for i in range(j):
            if keep_mask[i] and sim[i, j] >= threshold:
                keep_mask[j] = False
                logger.info(
                    "Dedup: dropping row %d (sim=%.3f vs row %d): %r",
                    j,
                    sim[i, j],
                    i,
                    questions[j][:80],
                )
                break

    df_clean = df[keep_mask].reset_index(drop=True)
    dropped = n - len(df_clean)
    logger.info("Dedup: kept %d / %d (dropped %d near-duplicates ≥ %.2f)", len(df_clean), n, dropped, threshold)
    return df_clean


# ---------------------------------------------------------------------------
# Step 8 — Export
# ---------------------------------------------------------------------------


_HOP_PREFIX_RE = re.compile(r"^<\d+-hop>\s*", re.MULTILINE)


def _build_chunk_crn_index(kg) -> list[tuple[str, str]]:
    """Return [(chunk_page_content, crn), ...] ordered by descending length.

    Walks the KG: for each CHUNK node, finds its parent DOCUMENT via the
    'child' relationship (source=doc, target=chunk) and reads
    document_metadata.crn. Longest chunks come first so substring lookups
    prefer more specific matches.
    """
    child_rels = [r for r in kg.relationships if r.type == "child"]
    chunk_to_doc = {r.target.id: r.source for r in child_rels}

    pairs: list[tuple[str, str]] = []
    for node in kg.nodes:
        if node.type.name != "CHUNK":
            continue
        content = node.properties.get("page_content", "")
        if not content:
            continue
        parent = chunk_to_doc.get(node.id)
        if not parent:
            continue
        crn = parent.properties.get("document_metadata", {}).get("crn", "") or ""
        pairs.append((content, crn))

    pairs.sort(key=lambda p: len(p[0]), reverse=True)
    return pairs


def _attribute_crns(ref_contexts, chunk_index: list[tuple[str, str]]) -> list[str]:
    """Return ALL CRNs whose source chunks appear in ref_contexts.

    Multi-hop questions span multiple chunks — each `reference_context` entry
    (after stripping `<N-hop>` markers) matches a different chunk. Walks every
    context and collects every distinct matching CRN. Returns CRNs sorted
    deterministically.
    """
    found: set[str] = set()
    if not ref_contexts or not isinstance(ref_contexts, list):
        return []
    for raw_ctx in ref_contexts:
        if not isinstance(raw_ctx, str):
            continue
        ctx = _HOP_PREFIX_RE.sub("", raw_ctx).strip()
        if not ctx:
            continue
        probe = ctx[:100]
        matched = False
        for chunk_content, crn in chunk_index:
            if probe in chunk_content:
                if crn:
                    found.add(crn)
                matched = True
                break
        if not matched and len(ctx) > 200:
            probe = ctx[100:200]
            for chunk_content, crn in chunk_index:
                if probe in chunk_content:
                    if crn:
                        found.add(crn)
                    break
    return sorted(found)


def _attribute_crn(ref_contexts, chunk_index: list[tuple[str, str]]) -> str:
    """Compatibility wrapper: returns first CRN as string (legacy single-CRN callers)."""
    crns = _attribute_crns(ref_contexts, chunk_index)
    return crns[0] if crns else ""


def _build_crn_to_course_id(kg) -> dict[str, str]:
    """Return ``{crn: course_id}`` from DOCUMENT-node metadata."""
    out: dict[str, str] = {}
    for node in kg.nodes:
        if node.type.name != "DOCUMENT":
            continue
        m = node.properties.get("document_metadata", {}) or {}
        crn = str(m.get("crn", "") or "")
        course_id = m.get("course_id", "") or ""
        if crn and course_id:
            out[crn] = course_id
    return out


def inject_course_codes(df, kg):
    """Prepend course identifier to single-syllabus stems that don't name one.

    Multi-CRN comparative items already name multiple courses, so they're
    left alone. For single-CRN items whose stem references generic
    "the course"/"the assignments"/"the final exam" without a dept-code
    (e.g. "ISEN 689"), prepend "In <COURSE_ID>, ...".
    """
    if "reference_contexts" not in df.columns or len(df) == 0:
        return df

    chunk_index = _build_chunk_crn_index(kg)
    crn_to_course = _build_crn_to_course_id(kg)

    new_questions: list[str] = []
    rewritten = 0
    for _, row in df.iterrows():
        question = str(row.get("user_input") or "")
        rc = row.get("reference_contexts")
        if isinstance(rc, str):
            try:
                rc = json.loads(rc)
            except json.JSONDecodeError:
                rc = [rc]
        crns = _attribute_crns(rc, chunk_index)

        if len(crns) == 1 and not _COURSE_CODE_RE.search(question) and _NEEDS_COURSE_CONTEXT_RE.search(question):
            course_id = crn_to_course.get(crns[0], "")
            if course_id:
                first = question[0] if question else ""
                if first.isupper() and not question[:4].isupper():
                    question = first.lower() + question[1:]
                question = f"In {course_id}, {question}"
                rewritten += 1

        new_questions.append(question)

    df = df.copy()
    df["user_input"] = new_questions
    logger.info("Course-code injection: rewrote %d / %d stems", rewritten, len(df))
    return df


def export_golden_set(df, documents: list[Document], output_path: Path, kg=None) -> None:
    """Export validated testset to XLSX matching golden_set schema.

    If ``kg`` is provided, CRN attribution uses a KG chunk → parent doc lookup
    (robust to Ragas-side header reformatting). Otherwise falls back to a
    legacy substring match against documents.
    """
    import openpyxl

    chunk_index: list[tuple[str, str]] | None = None
    if kg is not None:
        chunk_index = _build_chunk_crn_index(kg)
        logger.info("CRN attribution: built %d-entry chunk index from KG", len(chunk_index))

    # Legacy fallback path (kg=None)
    crn_lookup: dict[str, str] = {doc.metadata["source_file"]: doc.metadata.get("crn", "") for doc in documents}

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active

    columns = [
        "id",
        "question",
        "reference_answer",
        "expected_function",
        "human_notes",
        "reference_contexts",
        "crn",
    ]
    ws.append(columns)

    for i, (_, row) in enumerate(df.iterrows(), start=1):
        synthesizer_name = row.get("synthesizer_name", "")
        expected_function = SYNTHESIZER_TO_FUNCTION.get(synthesizer_name, "hybrid_course")

        ref_contexts = row.get("reference_contexts")
        if isinstance(ref_contexts, list):
            ref_contexts_str = json.dumps(ref_contexts, ensure_ascii=False)
        elif ref_contexts is not None:
            ref_contexts_str = str(ref_contexts)
        else:
            ref_contexts_str = ""

        if chunk_index is not None:
            crns_list = _attribute_crns(ref_contexts, chunk_index)
            crn = ", ".join(crns_list)
        else:
            crn = ""
            if ref_contexts and isinstance(ref_contexts, list):
                for ctx in ref_contexts:
                    if isinstance(ctx, str):
                        for src_file, src_crn in crn_lookup.items():
                            doc_content = next(
                                (d.page_content for d in documents if d.metadata["source_file"] == src_file),
                                "",
                            )
                            if ctx.strip()[:100] in doc_content:
                                crn = src_crn
                                break
                    if crn:
                        break

        ws.append(
            [
                i,
                row.get("user_input", ""),
                row.get("reference", ""),
                expected_function,
                "",
                ref_contexts_str,
                crn,
            ]
        )

    wb.save(output_path)
    logger.info("Golden set exported to %s (%d items)", output_path, len(df))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Generate RAGAS testset from syllabus eval corpus",
    )
    p.add_argument(
        "--corpus-dir",
        required=True,
        type=Path,
        help="Directory containing v3_step3_flat JSON files",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output XLSX path (default: tamu_data/evals/golden_sets/ragas_{date}.xlsx)",
    )
    # Legacy: kept for back-compat. If set and --target-size is the default, it
    # acts as the target. Otherwise --target-size wins.
    p.add_argument(
        "--testset-size",
        type=int,
        default=None,
        help="(Deprecated alias for --target-size)",
    )
    p.add_argument(
        "--target-size",
        type=int,
        default=20,
        help="Final validated item count desired (default: 20)",
    )
    p.add_argument(
        "--batch-size",
        type=int,
        default=5,
        help="Items generated per batch before validation+save (default: 5)",
    )
    p.add_argument(
        "--max-batches",
        type=int,
        default=10,
        help="Safety cap on number of generation batches (default: 10)",
    )
    p.add_argument(
        "--distribution",
        choices=list(DISTRIBUTION_PRESETS),
        default="default",
        help="Query distribution preset (default: 'default' = 50/30/20)",
    )
    p.add_argument(
        "--include-terms",
        type=str,
        default=None,
        help="Comma-separated term codes to include, e.g. 202611,202621,202641. If unset, all terms load.",
    )
    p.add_argument(
        "--kg-path",
        type=Path,
        default=Path("tamu_data/evals/ragas_kg.json"),
        help="KG cache file path (default: tamu_data/evals/ragas_kg.json)",
    )
    p.add_argument(
        "--rebuild-kg",
        action="store_true",
        help="Force KG rebuild even if cache exists",
    )
    p.add_argument(
        "--build-kg-only",
        action="store_true",
        help="Build/refresh KG, print stats, and exit before any question generation",
    )
    p.add_argument(
        "--provider",
        choices=["google", "tamu"],
        default="google",
        help="LLM provider (default: google)",
    )
    p.add_argument(
        "--temperature",
        type=float,
        default=0.4,
        help="LLM temperature (default: 0.4)",
    )
    p.add_argument(
        "--persona-file",
        type=str,
        default="tamu_data/evals/personas/course_shopping_student.yaml",
        help=(
            "Path to a persona YAML file (see src/tamubot/evals/personas.py). "
            "Pass an empty string to skip personas and fall back to Ragas defaults."
        ),
    )
    p.add_argument(
        "--keep-boilerplate-themes",
        action="store_true",
        help="Skip BOILERPLATE_THEMES filter on KG chunks (debug only)",
    )
    p.add_argument(
        "--seed-from",
        type=Path,
        default=None,
        help=(
            "Path to an existing golden-set XLSX (e.g. ragas_20260520_v3.xlsx). "
            "Items are loaded as the starting batch, and the CHUNK nodes that "
            "generated them are masked in the KG (themes/entities cleared) so "
            "Ragas doesn't re-draw the same scenarios."
        ),
    )
    p.add_argument(
        "--mask-whole-crn",
        action="store_true",
        help=(
            "When used with --seed-from, mask ALL chunks from any CRN already "
            "represented in the seed (not just chunks that produced seed items). "
            "Forces Ragas into completely uncovered courses."
        ),
    )
    p.add_argument(
        "--max-per-crn",
        type=int,
        default=3,
        help=(
            "Cap on items sharing the same source CRN. Ragas re-draws scenarios "
            "from the same chunks across batches; this prevents a single rich "
            "course from dominating the output. Default: 3"
        ),
    )
    p.add_argument(
        "--multi-hop-relation",
        type=str,
        default="entities_overlap_score",
        choices=["entities_overlap_score", "summary_similarity"],
        help=(
            "KG edge property the multi-hop-abstract synthesizer uses for clustering. "
            "Default 'entities_overlap_score' yields ~10 distinct topic clusters; "
            "'summary_similarity' tends to collapse to 2 hub-anchored templates "
            "because syllabi share too much structural language."
        ),
    )
    p.add_argument(
        "--debug-theme-matching",
        action="store_true",
        help="Log every ThemesPersonasMatching LLM call (themes in, mapping out)",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Load docs and print stats without generating",
    )
    return p.parse_args()


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(name)s %(levelname)s %(message)s",
    )

    args = parse_args()

    # Resolve target size: --target-size wins; --testset-size kept as legacy alias.
    target_size = args.target_size
    if args.testset_size is not None:
        target_size = args.testset_size
        print(f"[deprecation] --testset-size is an alias; using target_size={target_size}")

    include_terms: set[str] | None = None
    if args.include_terms:
        include_terms = {t.strip() for t in args.include_terms.split(",") if t.strip()}
        print(f"Filtering corpus to terms: {sorted(include_terms)}")

    # --- Load corpus ---
    documents = load_corpus(args.corpus_dir, include_terms=include_terms)
    print(f"\nCorpus: {len(documents)} documents from {args.corpus_dir}")

    if not documents:
        print("ERROR: No documents with chunks found. Check --corpus-dir path.")
        return

    token_counts = []
    try:
        from ragas.utils import num_tokens_from_string

        token_counts = [num_tokens_from_string(d.page_content) for d in documents]
        total = sum(token_counts)
        print(
            f"Token stats: total={total:,}, mean={total // len(documents):,}, "
            f"min={min(token_counts):,}, max={max(token_counts):,}"
        )
    except ImportError:
        print(f"Characters: total={sum(len(d.page_content) for d in documents):,}")

    for doc in documents:
        m = doc.metadata
        print(f"  {m['course_id']:>12s}  CRN={m['crn'] or '?':>6s}  {m['source_file']}")

    if args.dry_run:
        print("\n--dry-run: stopping before generation.")
        return

    # --- Build LLM + embeddings ---
    print(f"\nSetting up LLM (provider={args.provider}, temp={args.temperature})...")
    llm = build_llm(args.provider, args.temperature)

    print("Setting up embeddings (Google text-embedding-004)...")
    embedding_model = build_embeddings()

    # --- Build or load KG ---
    kg = build_or_load_kg(
        documents,
        llm,
        embedding_model,
        args.kg_path,
        args.rebuild_kg,
    )

    # Strip boilerplate themes (policy/admin) so multi-hop-abstract focuses on
    # course content. Done in-memory, KG cache file is untouched.
    if not args.keep_boilerplate_themes:
        filter_boilerplate_themes(kg)

    if args.debug_theme_matching:
        _install_theme_matching_logger()

    # KG stats summary
    try:
        n_nodes = len(kg.nodes)
        n_rels = len(kg.relationships)
        node_types: dict[str, int] = {}
        for n in kg.nodes:
            t = getattr(n.type, "value", str(n.type))
            node_types[t] = node_types.get(t, 0) + 1
        print(f"\nKG: nodes={n_nodes} ({node_types}), relationships={n_rels}")
    except Exception as e:  # noqa: BLE001
        print(f"(could not introspect KG: {e})")

    if args.build_kg_only:
        print(f"\n--build-kg-only: stopping after KG build. Cache at {args.kg_path}.")
        return

    # --- Query distribution ---
    query_distribution = build_query_distribution(
        llm, preset=args.distribution, multi_hop_relation=args.multi_hop_relation
    )
    print(f"Distribution preset: {args.distribution} (multi-hop relation: {args.multi_hop_relation})")

    # --- Load personas (optional; empty path = Ragas defaults) ---
    persona_list = None
    if args.persona_file:
        from tamubot.evals.personas import load_personas

        persona_list = load_personas(Path(args.persona_file))
        print(f"Loaded {len(persona_list)} persona(s): {[p.name for p in persona_list]}")
    else:
        print("No persona file provided — falling back to Ragas default persona.")

    # --- Batched generation loop ---
    import pandas as pd

    output_path = args.output or Path(f"tamu_data/evals/golden_sets/ragas_{datetime.now():%Y%m%d}.xlsx")

    # --- Optional seed from existing XLSX ---
    collected = pd.DataFrame()
    if args.seed_from is not None:
        import openpyxl

        wb = openpyxl.load_workbook(str(args.seed_from))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
        seed_df = pd.DataFrame(rows)
        # Normalize columns Ragas expects in the rest of the pipeline
        seed_df = seed_df.rename(columns={"question": "user_input", "reference_answer": "reference"})
        # Map expected_function back to synthesizer_name for dedup logging
        rev_map = {v: k for k, v in SYNTHESIZER_TO_FUNCTION.items()}
        seed_df["synthesizer_name"] = (
            seed_df.get("expected_function", "").map(rev_map).fillna("single_hop_specific_query_synthesizer")
        )

        # Reference_contexts stored as JSON strings — parse back to list
        def _parse_ctx(x):
            if isinstance(x, list):
                return x
            if not x:
                return []
            try:
                return json.loads(x)
            except json.JSONDecodeError, TypeError:
                return [x] if isinstance(x, str) else []

        seed_df["reference_contexts"] = seed_df["reference_contexts"].apply(_parse_ctx)
        print(f"Loaded {len(seed_df)} seed items from {args.seed_from}")
        mask_used_nodes(kg, seed_df, mask_whole_crn=args.mask_whole_crn)
        collected = seed_df

    for batch_num in range(1, args.max_batches + 1):
        if len(collected) >= target_size:
            break
        remaining = target_size - len(collected)
        # Ragas needs at least a few items per call to use multiple synthesizers
        this_batch = max(min(args.batch_size, remaining + 2), 3)
        print(
            f"\n[batch {batch_num}/{args.max_batches}] "
            f"generating {this_batch} (validated so far: {len(collected)}/{target_size})"
        )
        try:
            raw = generate_testset(
                kg,
                llm,
                embedding_model,
                query_distribution,
                this_batch,
                persona_list=persona_list,
            )
        except Exception as e:  # noqa: BLE001
            import traceback

            print(f"[batch {batch_num}] generation error: {e}")
            print(traceback.format_exc())
            continue

        validated = validate_testset(raw, documents, kg=kg)
        if len(validated) == 0:
            print(f"[batch {batch_num}] no items survived validation, continuing")
            continue

        validated = inject_course_codes(validated, kg)

        collected = pd.concat([collected, validated], ignore_index=True)
        collected = cap_per_crn(collected, kg, max_per_crn=args.max_per_crn)
        collected = deduplicate_by_question_similarity(collected, embedding_model)
        # Incremental save (trimmed to target_size if we overshot)
        export_golden_set(collected.head(target_size), documents, output_path, kg=kg)
        print(f"[batch {batch_num}] kept {len(validated)}, total {len(collected)} → saved {output_path}")

    if len(collected) == 0:
        print("ERROR: No items survived validation across all batches. Check KG quality.")
        return

    final = cap_per_crn(collected, kg, max_per_crn=args.max_per_crn)
    final = deduplicate_by_question_similarity(final, embedding_model).head(target_size)
    export_golden_set(final, documents, output_path, kg=kg)
    print(f"\nDone! Golden set: {output_path}  ({len(final)} items)")


if __name__ == "__main__":
    main()
