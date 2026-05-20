"""Unified eval runner — replaces run_benchmark.py and eval_chunking.py.

Generation is opt-in via --with-generation. Metrics are picked via --metrics
(comma-separated, or 'all'). A subset of questions can be re-run via --ids;
re-runs are linked to the SAME Langfuse dataset run (prior traces are
tagged 'superseded' but not deleted, since the SDK has no trace-delete primitive).

Usage:
    python -m tamubot.evals.run_eval \\
        --golden-set tamu_data/evals/golden_sets/<file>.xlsx \\
        --experiment <name> \\
        [--with-generation] \\
        [--ids 3,7,12] \\
        [--metrics faithfulness,context_recall] \\
        [--chunk-tag 300t_50o] [--chunks-collection chunks_eval] \\
        [--top-k N] [--threshold 0.35] \\
        [--capture-state] \\
        [--description "..."]
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

from tamubot.evals._runner_common import (  # noqa: E402
    filter_items_by_ids,
    find_existing_run_items,
    get_git_commit,
    load_dataset_item_ids,
    parse_id_filter,
    pre_arg,
    set_collection_env,
    stable_item_id,
    tag_trace_superseded,
    upsert_langfuse_dataset,
)

# Pre-parse collection / chunk-tag args BEFORE rag imports — these env vars
# must be set before rag/tools/mongo.py is imported.
set_collection_env(pre_arg("--chunks-collection"), pre_arg("--chunk-tag"))

from tamubot.evals.golden_set import append_run_column  # noqa: E402
from tamubot.evals.golden_set import load as load_golden_set
from tamubot.evals.run_benchmark import BenchmarkRow, write_excel, write_markdown  # noqa: E402
from tamubot.rag import RouterResult  # noqa: E402
from tamubot.rag.graph.pipeline import run_pipeline, run_pipeline_eval  # noqa: E402
from tamubot.rag.observability import (  # noqa: E402
    EvalInputs,
    available_metrics,
    eval_config,
    get_langfuse,
    resolve_metrics,
    run_evals,
    serialize_state,
    trace_context,
)

logger = logging.getLogger("tamubot.evals.run_eval")

STATE_DUMP_DIR = Path("tamu_data/evals/state_dumps")
NODE_CANDIDATES_DIR = Path("tamu_data/evals/golden_sets/_node_candidates")
REPORTS_DIR = Path("tamu_data/evals/reports")


# ---------------------------------------------------------------------------
# Default metric sets
# ---------------------------------------------------------------------------


def default_metrics(with_generation: bool) -> list[str]:
    if with_generation:
        return ["faithfulness", "answer_relevancy", "context_precision", "context_recall"]
    return ["context_precision", "context_recall"]


# ---------------------------------------------------------------------------
# Per-question execution
# ---------------------------------------------------------------------------


def _run_question(
    item: dict,
    *,
    with_generation: bool,
    metrics: list[str],
    experiment: str,
    run_name: str,
    capture_state: bool,
    question_idx: int,
) -> tuple[BenchmarkRow, dict[str, dict], Optional[str]]:
    """Execute one question. Returns (row, state_dump, trace_id)."""
    query = item["question"]
    expected_fn = str(item.get("expected_function", ""))

    obs = eval_config(
        experiment=experiment,
        run_name=run_name,
        with_generation=with_generation,
        metrics=metrics,
    )
    obs.metadata.update({"question_id": question_idx, "expected_function": expected_fn})

    t0 = time.perf_counter()
    error: Optional[str] = None
    chunks: list[dict] = []
    answer = ""
    rr = RouterResult(rewritten_query=query)
    timing_ms: dict = {}
    raw_state: dict = {}

    with trace_context(obs, query=query) as (lf_trace, trace_id):
        try:
            if with_generation:
                out = run_pipeline(query, return_timing=True, return_raw_state=True)
                chunks, rr_out, _gaps, _integrity, _conflicted, answer, timing_ms, raw_state = out
            else:
                chunks, rr_out, timing_ms, raw_state = run_pipeline_eval(query, return_raw_state=True)
            if rr_out is not None:
                rr = rr_out
            else:
                error = "pipeline returned no router_result"
        except Exception as e:
            error = f"pipeline: {e}"

        if lf_trace is not None:
            try:
                preview = answer[:500] if answer else f"{len(chunks)} chunks"
                lf_trace.update(output=preview)
            except Exception:
                pass

    total_ms = round((time.perf_counter() - t0) * 1000, 1)
    state_dump = serialize_state(raw_state) if (capture_state and raw_state) else {}

    # Metric evaluation
    faithfulness: Optional[float] = None
    answer_relevancy: Optional[float] = None
    metric_scores: dict[str, Optional[float]] = {}

    if metrics and chunks and not error:
        contexts = [c.get("content", "") for c in chunks if c.get("content")]
        reference = item.get("reference_answer") or ""
        try:
            scores = run_evals(
                obs,
                EvalInputs(
                    question=query,
                    contexts=contexts,
                    answer=answer,
                    reference=reference,
                    trace_id=trace_id,
                ),
            )
            faithfulness = scores.get("faithfulness")
            answer_relevancy = scores.get("answer_relevancy")
            metric_scores = dict(scores)
        except Exception as e:
            logger.warning(f"metrics failed for q{question_idx}: {e}")

    is_recurrent = rr.function == "recurrent"
    import re

    row = BenchmarkRow(
        question_id=question_idx,
        question=query,
        stratum=item.get("stratum", ""),
        source_course_id=str(item.get("source_course_id") or ""),
        expected_function=expected_fn,
        router_function=rr.function,
        router_function_correct=(rr.function == expected_fn),
        router_rewritten_query=rr.rewritten_query or "",
        router_course_ids=", ".join(rr.course_ids) if rr.course_ids else "",
        router_intent_type=rr.intent_type or "",
        chunks_retrieved=len(chunks),
        est_input_tokens=max(0, (len(query) + sum(len(c.get("content", "")) for c in chunks)) // 4),
        est_output_tokens=len(answer) // 4,
        pipeline_ms=round(total_ms - (timing_ms.get("generator_node") or 0.0), 1),
        generator_ms=timing_ms.get("generator_node") or 0.0,
        total_ms=total_ms,
        answer_full=answer,
        answer_preview=(answer[:120] + "…" if len(answer) > 120 else answer),
        citation_pass=bool(re.search(r"\[Source \d+\]", answer)),
        is_recurrent=is_recurrent,
        router_ms=timing_ms.get("router_node"),
        retrieval_ms=timing_ms.get("retrieval_node"),
        generator_node_ms=timing_ms.get("generator_node"),
        anchor_ms=timing_ms.get("anchor_node") if is_recurrent else None,
        eval_search_ms=timing_ms.get("eval_search_node") if is_recurrent else None,
        schedule_filter_ms=timing_ms.get("schedule_filter_node") if is_recurrent else None,
        merge_ms=timing_ms.get("merge_node") if is_recurrent else None,
        ragas_faithfulness=faithfulness,
        ragas_relevancy=answer_relevancy,
        metric_scores=metric_scores,
        error=error,
    )
    return row, state_dump, trace_id


# ---------------------------------------------------------------------------
# State capture sidecar + per-node candidate golden sets
# ---------------------------------------------------------------------------


def _write_sidecar_line(sidecar_path: Path, question_id: int, question: str, dump: dict[str, dict]) -> None:
    sidecar_path.parent.mkdir(parents=True, exist_ok=True)
    record = {"question_id": question_id, "question": question, "nodes": dump}
    with sidecar_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")


def _emit_node_candidate_xlsx(
    run_name: str,
    items: list[dict],
    rows: list[BenchmarkRow],
    dumps: list[dict[str, dict]],
) -> Optional[Path]:
    if not dumps or not any(dumps):
        return None
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed — skipping node candidate xlsx")
        return None

    out_dir = NODE_CANDIDATES_DIR / run_name
    out_dir.mkdir(parents=True, exist_ok=True)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    def _new_sheet(headers: list[str]):
        wb = openpyxl.Workbook()
        ws = wb.active
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
        ws.freeze_panes = "A2"
        return wb, ws

    router_wb, router_ws = _new_sheet(
        ["id", "question", "expected_function", "captured_function", "captured_course_ids", "captured_rewritten_query"]
    )
    retrieval_wb, retrieval_ws = _new_sheet(["id", "rewritten_query", "course_ids", "n_chunks", "captured_chunks_json"])
    generator_wb, generator_ws = _new_sheet(["id", "question", "n_chunks", "reference_answer", "captured_answer"])

    has_retrieval = False
    has_generator = False

    for row, dump, item in zip(rows, dumps, items):
        qid = row.question_id

        r = dump.get("router") or {}
        router_ws.append(
            [
                qid,
                row.question,
                row.expected_function,
                r.get("function", ""),
                ", ".join(r.get("course_ids", []) or []),
                r.get("rewritten_query", ""),
            ]
        )

        if "retrieval" in dump:
            has_retrieval = True
            rt = dump["retrieval"]
            retrieval_ws.append(
                [
                    qid,
                    rt.get("rewritten_query", ""),
                    ", ".join(rt.get("course_ids", []) or []),
                    rt.get("n_anchor", 0) + rt.get("n_followup", 0),
                    json.dumps(rt.get("chunks", []), ensure_ascii=False),
                ]
            )

        if "generator" in dump:
            has_generator = True
            g = dump["generator"]
            generator_ws.append(
                [
                    qid,
                    row.question,
                    g.get("n_chunks", 0),
                    item.get("reference_answer", "") or "",
                    g.get("answer", ""),
                ]
            )

    router_path = out_dir / "router.xlsx"
    router_wb.save(router_path)

    if has_retrieval:
        retrieval_wb.save(out_dir / "retrieval.xlsx")
    if has_generator:
        generator_wb.save(out_dir / "generator.xlsx")

    return out_dir


# ---------------------------------------------------------------------------
# Main loop
# ---------------------------------------------------------------------------


def _link_dataset_run_item(
    lf,
    *,
    dataset_name: str,
    run_name: str,
    run_description: Optional[str],
    dataset_item_id: str,
    trace_id: str,
    metadata: dict[str, Any],
) -> None:
    try:
        lf.api.dataset_run_items.create(
            run_name=run_name,
            run_description=run_description,
            dataset_item_id=dataset_item_id,
            trace_id=trace_id,
            metadata=metadata,
        )
    except Exception as e:
        logger.warning(f"dataset_run_items.create failed for item {dataset_item_id[:8]}: {e}")


def run(
    *,
    golden_path: Path,
    experiment: str,
    with_generation: bool,
    metrics: list[str],
    ids: Optional[set[int]],
    capture_state: bool,
    description: Optional[str],
    top_k: Optional[int],
    threshold: float,
    chunks_collection: Optional[str],
    chunk_tag: Optional[str],
    dataset_name: Optional[str],
) -> dict:
    """Run the eval. Returns a summary dict."""
    items = load_golden_set(golden_path)
    if not items:
        print(f"ERROR: golden set is empty or unreadable: {golden_path}")
        sys.exit(1)

    if ids is not None:
        items = filter_items_by_ids(items, ids)
        if not items:
            print(f"ERROR: --ids {sorted(ids)} matched no rows in {golden_path}")
            sys.exit(1)
        print(f"  --ids filter: running {len(items)} of {ids}")

    # Re-run mode: keep the existing dataset run name so new traces link under it.
    is_rerun = ids is not None
    run_name = experiment if is_rerun else f"{experiment}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    dataset_name = dataset_name or golden_path.stem

    lf = get_langfuse()
    print(
        f"  Langfuse: {'connected' if lf else 'not configured'}  |  "
        f"dataset='{dataset_name}'  run='{run_name}'  generation={with_generation}  metrics={metrics}"
    )

    prior_links: dict[str, str] = {}
    item_id_for_question: dict[str, str] = {}
    if lf:
        upsert_langfuse_dataset(lf, items if not is_rerun else load_golden_set(golden_path), dataset_name)
        item_id_for_question = load_dataset_item_ids(lf, dataset_name)
        if is_rerun:
            prior_links = find_existing_run_items(lf, dataset_name, run_name)

    # Output paths
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    sidecar_path = STATE_DUMP_DIR / f"{run_name}.jsonl"
    if capture_state and not is_rerun and sidecar_path.exists():
        sidecar_path.unlink()  # fresh full-run = fresh sidecar; reruns append

    rows: list[BenchmarkRow] = []
    dumps: list[dict[str, dict]] = []
    run_col: dict = {}

    for idx_in_filter, item in enumerate(items, start=1):
        qid_raw = item.get("id")
        question_id = int(qid_raw) if qid_raw is not None else idx_in_filter
        # Rate limit: 2s between questions (TAMU 30 RPM)
        if idx_in_filter > 1:
            time.sleep(2)
        print(f"  [{idx_in_filter}/{len(items)}] q{question_id}: {item['question'][:65]}")
        row, dump, trace_id = _run_question(
            item,
            with_generation=with_generation,
            metrics=metrics,
            experiment=experiment,
            run_name=run_name,
            capture_state=capture_state,
            question_idx=question_id,
        )
        rows.append(row)
        dumps.append(dump)

        if capture_state and dump:
            _write_sidecar_line(sidecar_path, question_id, item["question"], dump)

        # Build run-column string (kept compatible with eval_chunking format)
        if row.error:
            run_col[qid_raw] = f"ERROR: {row.error}"
        elif with_generation and row.answer_full:
            run_col[qid_raw] = row.answer_full
        else:
            # Retrieval-only: course §category score, score
            parts = []
            # We don't have raw chunks here — pull from state dump if present
            for c in (dump.get("retrieval", {}).get("chunks") if dump else None) or []:
                cid = c.get("course_id") or "?"
                cat = c.get("category") or "?"
                score = c.get("score")
                s = f"{score:.2f}" if isinstance(score, (int, float)) else "?"
                parts.append(f"{cid} §{cat} {s}")
            if parts:
                run_col[qid_raw] = ", ".join(parts)

        # Langfuse linking
        if lf and trace_id and item_id_for_question:
            item_id = item_id_for_question.get(item["question"]) or stable_item_id(item["question"])
            if is_rerun:
                prior_trace = prior_links.get(item_id)
                if prior_trace and prior_trace != trace_id:
                    tag_trace_superseded(lf, prior_trace)
            _link_dataset_run_item(
                lf,
                dataset_name=dataset_name,
                run_name=run_name,
                run_description=description,
                dataset_item_id=item_id,
                trace_id=trace_id,
                metadata={
                    "with_generation": with_generation,
                    "metrics": ",".join(metrics) if metrics else "",
                    "top_k": top_k if top_k is not None else "auto",
                    "threshold": threshold,
                    "chunks_collection": chunks_collection or "",
                    "chunk_tag": chunk_tag or "",
                    "rerun": is_rerun,
                },
            )

    if lf:
        try:
            lf.flush()
        except Exception:
            pass

    # Excel report (always full-pipeline shape; generation columns blank in retrieval-only mode)
    output_xlsx = REPORTS_DIR / f"eval_{run_name}.xlsx"
    write_excel(rows, run_name, str(golden_path), output_xlsx, do_ragas=bool(metrics))
    write_markdown(rows, run_name, output_xlsx.with_suffix(".md"))
    print(f"  Excel:  {output_xlsx}")

    # Per-node candidate golden sets
    node_dir = None
    if capture_state:
        node_dir = _emit_node_candidate_xlsx(run_name, items, rows, dumps)
        if node_dir:
            print(f"  Node candidates: {node_dir}")
        if sidecar_path.exists():
            print(f"  State sidecar: {sidecar_path}")

    # Golden set run column
    if run_col:
        try:
            append_run_column(golden_path, run_name, run_col)
            print(f"  Golden set run column appended: run:{run_name}")
        except Exception as e:
            logger.warning(f"append_run_column failed: {e}")

    # One score column per metric: run:<exp>:<metric>
    metric_names = sorted({k for r in rows for k in (r.metric_scores or {})})
    for m in metric_names:
        scores_by_id: dict = {}
        for r in rows:
            v = (r.metric_scores or {}).get(m)
            if isinstance(v, (int, float)):
                scores_by_id[r.question_id] = round(float(v), 4)
        if not scores_by_id:
            continue
        try:
            append_run_column(golden_path, f"{run_name}:{m}", scores_by_id)
            print(f"  Golden set score column appended: run:{run_name}:{m}")
        except Exception as e:
            logger.warning(f"append_run_column ({m}) failed: {e}")

    return {
        "run_name": run_name,
        "n_questions": len(rows),
        "n_errors": sum(1 for r in rows if r.error),
        "report": str(output_xlsx),
        "sidecar": str(sidecar_path) if (capture_state and sidecar_path.exists()) else None,
        "node_candidates_dir": str(node_dir) if node_dir else None,
        "git_commit": get_git_commit(),
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Unified TamuBot eval runner — retrieval-only by default, --with-generation for full pipeline."
    )
    p.add_argument("--golden-set", type=Path, required=False, help="Path to golden set .xlsx")
    p.add_argument("--experiment", required=False, help="Experiment name (also used as Langfuse run name for reruns)")
    p.add_argument("--with-generation", action="store_true", help="Run generator node + faithfulness/answer_relevancy")
    p.add_argument(
        "--ids",
        type=str,
        default=None,
        help="Comma-separated question ids to re-run (e.g. 3,7,12). When set, links to the existing run name.",
    )
    p.add_argument(
        "--metrics",
        type=str,
        default=None,
        help=f"Comma-separated metrics ({available_metrics() if False else 'see --list-metrics'}) or 'all'. "
        "Default depends on --with-generation.",
    )
    p.add_argument("--list-metrics", action="store_true", help="Print available metric names and exit")
    p.add_argument("--capture-state", action="store_true", help="Write per-node sidecar JSONL + candidate xlsx files")
    p.add_argument("--description", type=str, default=None, help="Human-readable run description (Langfuse)")
    p.add_argument("--top-k", type=int, default=None, help="Override rerank top_k")
    p.add_argument("--threshold", type=float, default=0.35, help="Voyage-3 relevance threshold (default 0.35)")
    p.add_argument("--chunks-collection", type=str, default=None, help="MongoDB chunks collection override")
    p.add_argument("--chunk-tag", type=str, default=None, help="Filter retrieval to chunks with this chunk_tag")
    p.add_argument("--dataset", type=str, default=None, help="Langfuse dataset name (default: golden-set stem)")
    return p.parse_args()


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    args = parse_args()

    if args.list_metrics:
        print("Available metrics:")
        for m in available_metrics():
            print(f"  - {m}")
        return

    if not args.golden_set or not args.experiment:
        print("ERROR: --golden-set and --experiment are required (omit only with --list-metrics)")
        sys.exit(2)

    if not args.golden_set.exists():
        print(f"ERROR: golden set not found: {args.golden_set}")
        sys.exit(1)

    metrics = resolve_metrics(args.metrics, default_metrics(args.with_generation))
    ids = parse_id_filter(args.ids)

    summary = run(
        golden_path=args.golden_set,
        experiment=args.experiment,
        with_generation=args.with_generation,
        metrics=metrics,
        ids=ids,
        capture_state=args.capture_state,
        description=args.description,
        top_k=args.top_k,
        threshold=args.threshold,
        chunks_collection=args.chunks_collection,
        chunk_tag=args.chunk_tag,
        dataset_name=args.dataset,
    )

    print("\n" + "=" * 60)
    print(f"  Run: {summary['run_name']}")
    print(f"  Questions: {summary['n_questions']}   Errors: {summary['n_errors']}")
    print(f"  Report: {summary['report']}")
    if summary.get("sidecar"):
        print(f"  Sidecar: {summary['sidecar']}")
    if summary.get("node_candidates_dir"):
        print(f"  Node candidates: {summary['node_candidates_dir']}")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
