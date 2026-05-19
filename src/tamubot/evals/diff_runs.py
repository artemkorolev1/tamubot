"""Compare two run:<experiment> columns from a golden-set xlsx side by side.

Reads two run columns from one golden set, writes an Excel report with the
answer / score / delta per question, with conditional fills for regression vs.
improvement on a chosen numeric metric.

Usage:
    python -m tamubot.evals.diff_runs \\
        --golden-set tamu_data/evals/golden_sets/ragas_20260520_v6.xlsx \\
        --left run:isen_v6_baseline \\
        --right run:isen_v6_fix \\
        --output tamu_data/evals/reports/diff_baseline_vs_fix.xlsx \\
        [--metric ragas_faithfulness]

Optional --metric reads numeric scores from a corresponding `<run>:<metric>`
column if present (e.g. produced by a future extension); if absent, the report
shows only the answers and a changed Y/N flag.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Optional


def _load_columns(path: Path, columns: list[str]) -> tuple[list[str], list[dict]]:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required.")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    missing = [c for c in columns if c not in headers]
    if missing:
        print(f"ERROR: golden set missing columns: {missing}. Available: {headers}")
        sys.exit(1)
    idxs = {c: headers.index(c) for c in columns}

    rows: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append({c: r[idxs[c]] for c in columns})
    wb.close()
    return headers, rows


def _coerce_num(v) -> Optional[float]:
    try:
        return float(v) if v is not None and v != "" else None
    except TypeError, ValueError:
        return None


def _stringify(v) -> str:
    return "" if v is None else str(v)


def write_diff(
    golden_path: Path,
    left: str,
    right: str,
    output: Path,
    metric: Optional[str] = None,
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("ERROR: openpyxl required.")
        sys.exit(1)

    needed = ["id", "question", left, right]
    metric_left_col = f"{left}:{metric}" if metric else None
    metric_right_col = f"{right}:{metric}" if metric else None
    optional = [c for c in (metric_left_col, metric_right_col) if c]

    headers, rows_in = _load_columns(golden_path, needed)
    # Fetch optional metric columns separately so absence is non-fatal.
    have_metric = bool(metric) and metric_left_col in headers and metric_right_col in headers
    if metric and not have_metric:
        print(f"  Note: metric columns {metric_left_col}/{metric_right_col} not found — emitting text-only diff.")

    if have_metric:
        _, metric_rows = _load_columns(golden_path, ["id"] + optional)
        metric_by_id = {m["id"]: m for m in metric_rows}
    else:
        metric_by_id = {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Diff"

    out_headers = ["id", "question", "left_answer", "right_answer", "changed"]
    if have_metric:
        out_headers.extend(["left_score", "right_score", "delta"])

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    fill_better = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_worse = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for i, h in enumerate(out_headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill

    widths = [6, 50, 60, 60, 8] + ([12, 12, 10] if have_metric else [])
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    n_changed = 0
    n_better = 0
    n_worse = 0

    for row_idx, r in enumerate(rows_in, start=2):
        left_ans = _stringify(r.get(left))
        right_ans = _stringify(r.get(right))
        changed = left_ans != right_ans
        if changed:
            n_changed += 1

        ws.cell(row=row_idx, column=1, value=r["id"])
        ws.cell(row=row_idx, column=2, value=r["question"]).alignment = wrap
        ws.cell(row=row_idx, column=3, value=left_ans).alignment = wrap
        ws.cell(row=row_idx, column=4, value=right_ans).alignment = wrap
        ws.cell(row=row_idx, column=5, value="Y" if changed else "")

        if have_metric:
            m = metric_by_id.get(r["id"], {})
            ls = _coerce_num(m.get(metric_left_col))
            rs = _coerce_num(m.get(metric_right_col))
            delta = (rs - ls) if (ls is not None and rs is not None) else None
            ws.cell(row=row_idx, column=6, value=ls)
            ws.cell(row=row_idx, column=7, value=rs)
            ws.cell(row=row_idx, column=8, value=round(delta, 4) if delta is not None else None)
            if delta is not None:
                fill = fill_better if delta > 0 else (fill_worse if delta < 0 else None)
                if fill:
                    for c_idx in range(6, 9):
                        ws.cell(row=row_idx, column=c_idx).fill = fill
                if delta > 0:
                    n_better += 1
                elif delta < 0:
                    n_worse += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary tab
    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum["A1"] = "Diff Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)
    info = [
        ("Golden set", str(golden_path)),
        ("Left column", left),
        ("Right column", right),
        ("Metric column", metric or "(none)"),
        ("Rows compared", len(rows_in)),
        ("Rows changed", n_changed),
    ]
    if have_metric:
        info.extend([("Improved", n_better), ("Regressed", n_worse)])
    for i, (k, v) in enumerate(info, start=3):
        ws_sum.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws_sum.cell(row=i, column=2, value=v)
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 60

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"  Diff written: {output}  ({n_changed}/{len(rows_in)} changed)")


def main() -> None:
    p = argparse.ArgumentParser(description="Compare two run:* columns from a golden-set xlsx.")
    p.add_argument("--golden-set", type=Path, required=True)
    p.add_argument("--left", type=str, required=True, help="Left run column name (e.g. run:exp_a)")
    p.add_argument("--right", type=str, required=True, help="Right run column name (e.g. run:exp_b)")
    p.add_argument("--output", type=Path, required=True, help="Output .xlsx path")
    p.add_argument(
        "--metric",
        type=str,
        default=None,
        help="Optional numeric metric (e.g. ragas_faithfulness). Expects columns "
        "'<left>:<metric>' and '<right>:<metric>'; absent → text-only diff.",
    )
    args = p.parse_args()

    if not args.golden_set.exists():
        print(f"ERROR: golden set not found: {args.golden_set}")
        sys.exit(1)

    write_diff(args.golden_set, args.left, args.right, args.output, args.metric)


if __name__ == "__main__":
    main()
