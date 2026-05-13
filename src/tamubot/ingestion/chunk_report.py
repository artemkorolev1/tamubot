"""
tamubot.ingestion.chunk_report

Generate an Excel report for v4 semantic chunking results.

Usage:
    # Test on subset:
    python -m tamubot.ingestion.chunk_report \
        --input-dir data/syllabi/silver/04_hierarchy \
        --enrichment-dir data/syllabi/silver/05_enrich \
        --output data/syllabi/silver/06_chunk/chunk_report.xlsx \
        --files 202641_ISEN_613_601_34731,202641_ISEN_691_603_15312

    # All files:
    python -m tamubot.ingestion.chunk_report \
        --input-dir data/syllabi/silver/04_hierarchy \
        --enrichment-dir data/syllabi/silver/05_enrich \
        --output data/syllabi/silver/06_chunk/chunk_report.xlsx
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tamubot.ingestion.chunker_v4 import _tokens_approx, chunk_semantic

# ── Styling ─────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_RED_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")

# ── Summary sheet columns ────────────────────────────────────────────────────

SUMMARY_COLS = [
    "File",
    "Course",
    "Section",
    "Instructor",
    "Has Enrichment",
    "File Tokens",
    "Chunk Count",
    "Total Chunk Tok",
    "Dropped Meta",
    "Dropped Tiny",
    "Merged Tiny",
    "Preservation %",
    "Min Tok",
    "Max Tok",
    "Avg Tok",
    "Sections",
    "Kept",
    "Flagged",
    "Has Table",
    "Flags",
]

DETAIL_COLS = [
    "File",
    "Course",
    "Chunk Index",
    "Page",
    "Token Count",
    "Header Path",
    "Has Table",
    "Flags",
    "Content Preview",
]

ERROR_COLS = [
    "File",
    "Error",
]


# ── Data collection ─────────────────────────────────────────────────────────


def _load_enrichment_full(enrichment_dir: Path, stem: str) -> dict:
    """Load full enrichment JSON for a file stem, return empty dict if missing."""
    path = enrichment_dir / f"{stem}.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _compute_flags(
    file_tok: int,
    total_sections: int,
    flagged_count: int,
    has_enrichment: bool,
) -> str:
    flags = []
    if file_tok > 10000:
        flags.append("MASSIVE_FILE")
    if file_tok < 300:
        flags.append("TINY")
    if total_sections == 0:
        flags.append("ZERO_HEADERS")
    if flagged_count > 0:
        flags.append(f"OVERSIZED({flagged_count})")
    if not has_enrichment:
        flags.append("NO_ENRICHMENT")
    return ", ".join(flags) if flags else ""


def collect_chunk_data(
    input_dir: Path,
    enrichment_dir: Path,
    flag_threshold: int = 500,
    min_chunk_tokens: int = 15,
    file_stems: list[str] | None = None,
) -> tuple[list[dict], list[dict], list[dict]]:
    """Run semantic chunker on all markdown files and collect summary + detail + error rows."""
    md_files = sorted(input_dir.glob("*.md"))
    if file_stems:
        stem_set = set(file_stems)
        md_files = [f for f in md_files if f.stem in stem_set]

    summary_rows: list[dict] = []
    detail_rows: list[dict] = []
    error_rows: list[dict] = []

    for md_file in md_files:
        stem = md_file.stem

        try:
            markdown = md_file.read_text(encoding="utf-8")
            file_tok = _tokens_approx(markdown.strip())

            chunks, log = chunk_semantic(
                markdown,
                flag_threshold=flag_threshold,
                min_chunk_tokens=min_chunk_tokens,
            )

            enrichment = _load_enrichment_full(enrichment_dir, stem)
            meta = enrichment.get("course_metadata", {})
            has_enrichment = bool(meta)
            course_id = meta.get("course_id", "")
            section = meta.get("section", "")
            instructor = ""
            if meta.get("instructor"):
                instructor = meta["instructor"].get("name", "")

            # Resolve page numbers from enrichment headers
            from tamubot.ingestion.pipeline_v4 import _build_page_lookup, _resolve_page

            page_lookup = _build_page_lookup(enrichment)
            for chunk in chunks:
                chunk["page"] = _resolve_page(chunk, page_lookup)
                if chunk["page"] is None and not chunk["header_path"]:
                    chunk["page"] = 1

            sizes = [c["token_count"] for c in chunks] if chunks else [0]
            total_chunk_tok = sum(sizes)
            flagged_chunks = [c for c in chunks if "OVERSIZED" in c.get("flags", [])]
            table_chunks = [c for c in chunks if c["has_table"]]

            preservation = round(total_chunk_tok / file_tok * 100, 1) if file_tok > 0 else 0

            flags = _compute_flags(
                file_tok,
                log.get("total_sections", 0),
                len(flagged_chunks),
                has_enrichment,
            )

            summary_rows.append(
                {
                    "File": stem,
                    "Course": course_id,
                    "Section": section,
                    "Instructor": instructor,
                    "Has Enrichment": "yes" if has_enrichment else "no",
                    "File Tokens": file_tok,
                    "Chunk Count": len(chunks),
                    "Total Chunk Tok": total_chunk_tok,
                    "Dropped Meta": log.get("dropped_metadata", 0),
                    "Dropped Tiny": log.get("dropped_tiny", 0),
                    "Merged Tiny": log.get("merged_tiny", 0),
                    "Preservation %": preservation,
                    "Min Tok": min(sizes),
                    "Max Tok": max(sizes),
                    "Avg Tok": round(sum(sizes) / len(sizes)) if sizes else 0,
                    "Sections": log.get("total_sections", 0),
                    "Kept": log.get("kept", 0),
                    "Flagged": len(flagged_chunks),
                    "Has Table": len(table_chunks),
                    "Flags": flags,
                }
            )

            for chunk in chunks:
                preview = chunk["content"][:120].replace("\n", " ")
                detail_rows.append(
                    {
                        "File": stem,
                        "Course": course_id,
                        "Chunk Index": chunk["chunk_index"],
                        "Page": chunk.get("page") or "",
                        "Token Count": chunk["token_count"],
                        "Header Path": chunk["header_path"],
                        "Has Table": "yes" if chunk["has_table"] else "",
                        "Flags": ", ".join(chunk.get("flags", [])),
                        "Content Preview": preview,
                    }
                )

        except Exception as exc:
            error_rows.append({"File": stem, "Error": str(exc)})

    return summary_rows, detail_rows, error_rows


# ── Excel writing ───────────────────────────────────────────────────────────


def _write_header_row(ws, cols: list[str], row: int = 1) -> None:
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _auto_width(ws, cols: list[str], max_width: int = 40) -> None:
    for col_idx, col_name in enumerate(cols, 1):
        width = min(max(len(col_name) + 2, 10), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def generate_chunk_report(
    input_dir: Path,
    enrichment_dir: Path,
    output_path: Path,
    flag_threshold: int = 500,
    min_chunk_tokens: int = 15,
    file_stems: list[str] | None = None,
) -> Path:
    """Generate the semantic chunking Excel report."""
    summary_rows, detail_rows, error_rows = collect_chunk_data(
        input_dir,
        enrichment_dir,
        flag_threshold=flag_threshold,
        min_chunk_tokens=min_chunk_tokens,
        file_stems=file_stems,
    )

    wb = Workbook()

    # ── Sheet 1: Chunking summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_header_row(ws_summary, SUMMARY_COLS)
    _auto_width(ws_summary, SUMMARY_COLS)

    for row_idx, row_data in enumerate(summary_rows, 2):
        for col_idx, col_name in enumerate(SUMMARY_COLS, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
            # Highlight flags
            if col_name == "Flags" and row_data.get("Flags"):
                if "OVERSIZED" in row_data["Flags"] or "ZERO_HEADERS" in row_data["Flags"]:
                    cell.fill = _RED_FILL
                elif "NO_ENRICHMENT" in row_data["Flags"]:
                    cell.fill = _YELLOW_FILL
                else:
                    cell.fill = _YELLOW_FILL
            if col_name == "Flagged" and row_data.get("Flagged", 0) > 0:
                cell.fill = _YELLOW_FILL
            if col_name == "Has Enrichment" and row_data.get("Has Enrichment") == "no":
                cell.fill = _YELLOW_FILL

    ws_summary.freeze_panes = "A2"

    # ── Sheet 2: Chunk detail ──
    ws_detail = wb.create_sheet("Chunk Detail")
    _write_header_row(ws_detail, DETAIL_COLS)
    _auto_width(ws_detail, DETAIL_COLS)
    ws_detail.column_dimensions["I"].width = 60

    for row_idx, row_data in enumerate(detail_rows, 2):
        for col_idx, col_name in enumerate(DETAIL_COLS, 1):
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
            if col_name == "Content Preview":
                cell.alignment = _WRAP
            if col_name == "Token Count" and row_data.get("Token Count", 0) > flag_threshold:
                cell.fill = _YELLOW_FILL
            if col_name == "Flags" and row_data.get("Flags"):
                if "OVERSIZED" in row_data["Flags"]:
                    cell.fill = _YELLOW_FILL

    ws_detail.freeze_panes = "A2"

    # ── Sheet 3: Errors ──
    ws_errors = wb.create_sheet("Errors")
    _write_header_row(ws_errors, ERROR_COLS)
    _auto_width(ws_errors, ERROR_COLS, max_width=80)

    for row_idx, row_data in enumerate(error_rows, 2):
        for col_idx, col_name in enumerate(ERROR_COLS, 1):
            cell = ws_errors.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
            cell.fill = _RED_FILL

    ws_errors.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    # Print summary
    total_files = len(summary_rows)
    total_chunks = sum(r["Chunk Count"] for r in summary_rows)
    total_flagged = sum(r["Flagged"] for r in summary_rows)
    total_dropped_meta = sum(r["Dropped Meta"] for r in summary_rows)
    total_dropped_tiny = sum(r["Dropped Tiny"] for r in summary_rows)
    total_merged_tiny = sum(r["Merged Tiny"] for r in summary_rows)
    flagged_files = sum(1 for r in summary_rows if r["Flags"])
    print(f"\nChunk Report: {output_path}")
    print(f"  Files: {total_files}, Errors: {len(error_rows)}")
    print(f"  Total chunks: {total_chunks}")
    print(f"  Dropped: {total_dropped_meta} metadata, {total_dropped_tiny} tiny, {total_merged_tiny} merged")
    print(f"  Flagged chunks (>{flag_threshold} tok): {total_flagged}")
    print(f"  Flagged files: {flagged_files}")
    print(f"  Config: flag_threshold={flag_threshold}, min_tok={min_chunk_tokens}")

    return output_path


# ── CLI ──────────────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate chunking report for v4 pipeline")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("data/syllabi/silver/04_hierarchy"),
        help="Directory with hierarchy-corrected markdown files",
    )
    parser.add_argument(
        "--enrichment-dir",
        type=Path,
        default=Path("data/syllabi/silver/05_enrich"),
        help="Directory with enrichment JSON files",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/syllabi/silver/06_chunk/chunk_report.xlsx"),
        help="Output Excel path",
    )
    parser.add_argument("--flag-threshold", type=int, default=500)
    parser.add_argument("--min-chunk-tokens", type=int, default=30)
    parser.add_argument(
        "--files",
        default=None,
        help="Comma-separated file stems to process (default: all)",
    )

    args = parser.parse_args()
    file_stems = [s.strip() for s in args.files.split(",")] if args.files else None

    generate_chunk_report(
        input_dir=args.input_dir,
        enrichment_dir=args.enrichment_dir,
        output_path=args.output,
        flag_threshold=args.flag_threshold,
        min_chunk_tokens=args.min_chunk_tokens,
        file_stems=file_stems,
    )


if __name__ == "__main__":
    main()
