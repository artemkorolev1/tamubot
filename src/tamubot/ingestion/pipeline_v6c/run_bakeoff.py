"""Bake-off runner: run opendataloader-pdf over the same 10 STAT stems as v6
and v6b, write bronze artifacts under v6c/bronze/, and emit pilot_bakeoff.xlsx
with per-stem timing/header counts and a side-by-side header count vs v6/v6b.

Usage:
    python -m tamubot.ingestion.pipeline_v6c.run_bakeoff [STAT]
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from tamubot.ingestion.pipeline_v5.schemas import HeadersSidecar
from tamubot.ingestion.pipeline_v5.util import DATA_ROOT, hash_file
from tamubot.ingestion.pipeline_v6c import paths as v6c_paths
from tamubot.ingestion.pipeline_v6c.assets.bronze_odl import odl_convert

_HEADER_RE = re.compile(r"^#{1,6}\s+", re.MULTILINE)

# Letter-drop dictionary from the bake-off investigation. These are
# known-buggy artifacts produced by veraPDF's text extraction on
# Simple Syllabus PDFs. Used as a soft-signal column in the report.
_LETTER_DROP_TOKENS = (
    r"\bColege\b",
    r"\bSylabus\b",
    r"\bMeting\b",
    r"\bAditional\b",
    r"\bTextbok\b",
    r"\bwil\b",
    r"\bfal\b",
    r"\binstaling\b",
)
_LETTER_DROP_RE = re.compile("|".join(_LETTER_DROP_TOKENS))


def _stems_for(dept: str) -> list[str]:
    """Source of truth: stems present under <dept>/v6/bronze/ (v6's bake-off set)."""
    v6_bronze = DATA_ROOT / dept.upper() / "v6" / "bronze"
    if not v6_bronze.exists():
        raise FileNotFoundError(f"v6 bronze dir missing for {dept!r}: {v6_bronze}")
    return sorted(p.stem for p in v6_bronze.glob("*.md"))


def _peer_md(stem: str, peer: str) -> Path | None:
    """Locate the bronze markdown for a peer pipeline (v5, v6, v6b). None if absent."""
    p = DATA_ROOT / stem.split("_")[1].upper() / peer / "bronze" / f"{stem}.md"
    return p if p.exists() else None


def _count_headers(md_path: Path | None) -> int | None:
    if md_path is None:
        return None
    return len(_HEADER_RE.findall(md_path.read_text(encoding="utf-8")))


def _convert_one(stem: str) -> dict:
    pdf = v6c_paths.raw_path(stem)
    bronze_dir = v6c_paths.bronze_md_path(stem).parent
    bronze_dir.mkdir(parents=True, exist_ok=True)

    result = odl_convert(pdf, output_dir=bronze_dir)

    md_out = v6c_paths.bronze_md_path(stem)
    md_out.write_text(result.markdown, encoding="utf-8")

    sidecar_out = v6c_paths.bronze_sidecar_path(stem)
    sidecar = HeadersSidecar(headers=result.headers)
    sidecar_out.write_text(sidecar.model_dump_json(indent=2), encoding="utf-8")

    odl_json_out = v6c_paths.bronze_odl_json_path(stem)
    odl_json_out.write_text(json.dumps(result.raw_json, indent=2, ensure_ascii=False), encoding="utf-8")

    drops = len(_LETTER_DROP_RE.findall(result.markdown))

    return {
        "stem": stem,
        "is_hp": stem.endswith("_HP"),
        "page_count": result.page_count,
        "v6c_headers": len(result.headers),
        "v6c_chars": len(result.markdown),
        "v6c_tokens_est": len(result.markdown) // 4,
        "v6c_timing_s": round(result.timing_s, 2),
        "letter_drop_hits": drops,
        "repair_word_fixes": result.repair_word_fixes,
        "repair_merge_fixes": result.repair_merge_fixes,
        "v6c_md_sha": hash_file(md_out),
        "v5_headers": _count_headers(_peer_md(stem, "v5")),
        "v6_headers": _count_headers(_peer_md(stem, "v6")),
        "v6b_headers": _count_headers(_peer_md(stem, "v6b")),
    }


def _write_report(dept: str, rows: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "v6c bake-off"

    headers = [
        "stem",
        "source",
        "pages",
        "v6c_headers",
        "v5_headers",
        "v6_headers",
        "v6b_headers",
        "v6c_chars",
        "v6c_tokens_est",
        "v6c_timing_s",
        "letter_drop_hits",
        "repair_word_fixes",
        "repair_merge_fixes",
        "v6c_md_sha",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

    flag_fill = PatternFill("solid", fgColor="FFD9D9")
    for r in rows:
        ws.append(
            [
                r["stem"],
                "HowdyPortal" if r["is_hp"] else "Simple Syllabus",
                r["page_count"],
                r["v6c_headers"],
                r["v5_headers"],
                r["v6_headers"],
                r["v6b_headers"],
                r["v6c_chars"],
                r["v6c_tokens_est"],
                r["v6c_timing_s"],
                r["letter_drop_hits"],
                r["repair_word_fixes"],
                r["repair_merge_fixes"],
                r["v6c_md_sha"],
            ]
        )
        if r["letter_drop_hits"] > 0:
            for cell in ws[ws.max_row]:
                cell.fill = flag_fill

    for col_idx, col_name in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = max(12, min(40, len(col_name) + 2))
    ws.column_dimensions["A"].width = 40

    out = v6c_paths.bakeoff_report_path(dept)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main(argv: list[str]) -> int:
    dept = (argv[1] if len(argv) > 1 else "STAT").upper()
    stems = _stems_for(dept)
    print(f"v6c bake-off on {dept}: {len(stems)} stems")

    rows: list[dict] = []
    for i, stem in enumerate(stems, 1):
        print(f"  [{i:>2}/{len(stems)}] {stem}", flush=True)
        rows.append(_convert_one(stem))

    out = _write_report(dept, rows)
    print(f"\nReport: {out}")
    n_hp = sum(1 for r in rows if r["is_hp"])
    n_ss = len(rows) - n_hp
    drop_total = sum(r["letter_drop_hits"] for r in rows)
    drop_hp = sum(r["letter_drop_hits"] for r in rows if r["is_hp"])
    drop_ss = drop_total - drop_hp
    print(f"  HP files: {n_hp}, letter-drop hits: {drop_hp}")
    print(f"  SS files: {n_ss}, letter-drop hits: {drop_ss}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
