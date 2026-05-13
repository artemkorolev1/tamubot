"""Incremental Excel report writer for the Docling v4 pipeline.

After each file is processed by any pipeline step, call the appropriate
``update_*`` function to write results to the pipeline Excel report
immediately.  This ensures partial progress is preserved if a long-running
step (e.g. image recovery, validation) is interrupted.

Usage from filters / validators:
    from tamubot.ingestion.report_writer import get_report, update_image_recovery

    report = get_report()  # returns Path or None
    if report:
        update_image_recovery(report, stem, before, after, recovered)
"""

from __future__ import annotations

from pathlib import Path

DEFAULT_REPORT_PATH = Path("data/syllabi/silver/docling_pipeline_v4_report.xlsx")


def get_report(report_path: Path | None = None) -> Path | None:
    """Return the report path if it exists, else None."""
    path = report_path or DEFAULT_REPORT_PATH
    return path if path.exists() else None


# ── Shared helpers ────────────────────────────────────────────────────────────


def _load_workbook(report_path: Path):
    """Load workbook + build column map for Summary sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(report_path)
    ws = wb["Summary"]
    col_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            col_map[val] = c
    return wb, ws, col_map


def _find_row(ws, file_stem: str) -> int | None:
    """Find the row number for a given file stem in column 1."""
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == file_stem:
            return r
    return None


def _ensure_column(ws, col_map: dict[str, int], header: str) -> int:
    """Create a column if it doesn't exist yet, return its index."""
    if header in col_map:
        return col_map[header]
    import openpyxl

    new_col = ws.max_column + 1
    ws.cell(1, new_col, header).font = openpyxl.styles.Font(bold=True)
    col_map[header] = new_col
    return new_col


def _save(wb, report_path: Path) -> None:
    """Save workbook, tolerating permission errors."""
    try:
        wb.save(report_path)
    except PermissionError:
        print(f"    WARN: Report locked, could not save: {report_path}")


# ── Image recovery ────────────────────────────────────────────────────────────


def update_image_recovery(
    report_path: Path,
    file_stem: str,
    markers_before: int,
    markers_after: int,
    recovered: bool,
) -> None:
    """Write image recovery results for one file to the report."""
    try:
        from openpyxl.styles import PatternFill

        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        wb, ws, col_map = _load_workbook(report_path)
        row = _find_row(ws, file_stem)
        if row is None:
            return

        c_before = _ensure_column(ws, col_map, "Img Markers Before")
        c_after = _ensure_column(ws, col_map, "Img Markers After")
        c_flag = _ensure_column(ws, col_map, "Gemini Recovered")

        ws.cell(row, c_before, markers_before)
        ws.cell(row, c_after, markers_after)
        cell = ws.cell(row, c_flag)
        cell.value = "YES" if recovered else "FAILED"
        cell.fill = green if recovered else red

        _save(wb, report_path)
    except Exception as e:
        print(f"    WARN: Could not update report for {file_stem}: {e}")


# ── False positive ────────────────────────────────────────────────────────────


def update_false_positive(
    report_path: Path,
    file_stem: str,
    demoted_count: int,
    cleanups_applied: int,
    demoted_headers: list[str],
) -> None:
    """Write false-positive filter results for one file to the report."""
    try:
        wb, ws, col_map = _load_workbook(report_path)
        row = _find_row(ws, file_stem)
        if row is None:
            return

        c_demoted = _ensure_column(ws, col_map, "FP Headers Demoted")
        c_cleanups = _ensure_column(ws, col_map, "FP Cleanups")

        ws.cell(row, c_demoted, demoted_count)
        ws.cell(row, c_cleanups, cleanups_applied)

        _save(wb, report_path)
    except Exception as e:
        print(f"    WARN: Could not update report for {file_stem}: {e}")


# ── Boilerplate ───────────────────────────────────────────────────────────────


def update_boilerplate(
    report_path: Path,
    file_stem: str,
    sections_stripped: int,
    tokens_in: int,
    tokens_out: int,
    stripped_headers: list[str],
) -> None:
    """Write boilerplate filter results for one file to the report."""
    try:
        wb, ws, col_map = _load_workbook(report_path)
        row = _find_row(ws, file_stem)
        if row is None:
            return

        c_stripped = _ensure_column(ws, col_map, "BP Sections Stripped")
        c_tokens_in = _ensure_column(ws, col_map, "BP Tokens In")
        c_tokens_out = _ensure_column(ws, col_map, "BP Tokens Out")

        ws.cell(row, c_stripped, sections_stripped)
        ws.cell(row, c_tokens_in, tokens_in)
        ws.cell(row, c_tokens_out, tokens_out)

        _save(wb, report_path)
    except Exception as e:
        print(f"    WARN: Could not update report for {file_stem}: {e}")


# ── Hierarchy ─────────────────────────────────────────────────────────────────


def update_hierarchy(
    report_path: Path,
    file_stem: str,
    corrected: bool,
    levels_before: dict[int, int],
    levels_after: dict[int, int],
) -> None:
    """Write hierarchy filter results for one file to the report."""
    try:
        wb, ws, col_map = _load_workbook(report_path)
        row = _find_row(ws, file_stem)
        if row is None:
            return

        c_corrected = _ensure_column(ws, col_map, "Hierarchy Corrected")
        c_levels = _ensure_column(ws, col_map, "Heading Levels")

        ws.cell(row, c_corrected, "YES" if corrected else "—")
        levels_str = ", ".join(f"H{k}={v}" for k, v in sorted(levels_after.items()))
        ws.cell(row, c_levels, levels_str)

        _save(wb, report_path)
    except Exception as e:
        print(f"    WARN: Could not update report for {file_stem}: {e}")


# ── Validation ────────────────────────────────────────────────────────────────


def update_validation(
    report_path: Path,
    file_stem: str,
    version_label: str,
    counts: dict[str, int],
    findings: dict[str, list[str]],
) -> None:
    """Write validation results for one file to the report.

    Updates both the Summary sheet (count columns) and the per-category
    detail sheets (Content Pres., Strip Compl., Structural, Metadata).
    """
    try:
        wb, ws, col_map = _load_workbook(report_path)
        row = _find_row(ws, file_stem)
        if row is None:
            return

        # Summary columns: "{version_label} Content", etc.
        cat_short = {
            "content_preservation": "Content",
            "strip_completeness": "Strip",
            "structural_coherence": "Structural",
            "metadata_accuracy": "Metadata",
        }
        total = 0
        for cat, short in cat_short.items():
            col_header = f"{version_label} {short}"
            col = _ensure_column(ws, col_map, col_header)
            count = counts.get(cat, 0)
            ws.cell(row, col, count)
            total += count

        col_total = _ensure_column(ws, col_map, f"{version_label} Total")
        ws.cell(row, col_total, total)

        # Detail sheets
        sheet_map = {
            "content_preservation": "Content Pres.",
            "strip_completeness": "Strip Compl.",
            "structural_coherence": "Structural",
            "metadata_accuracy": "Metadata",
        }
        for cat, sheet_name in sheet_map.items():
            if sheet_name not in wb.sheetnames:
                continue
            detail_ws = wb[sheet_name]

            # Find the file row in the detail sheet
            detail_row = None
            for r in range(2, detail_ws.max_row + 1):
                if detail_ws.cell(r, 1).value == file_stem:
                    detail_row = r
                    break
            if detail_row is None:
                continue

            # Find or create the version columns
            detail_col_map: dict[str, int] = {}
            for c in range(1, detail_ws.max_column + 1):
                val = detail_ws.cell(1, c).value
                if val:
                    detail_col_map[val] = c

            count_header = f"{version_label} Count"
            findings_header = f"{version_label} Findings"

            if count_header not in detail_col_map:
                import openpyxl

                nc = detail_ws.max_column + 1
                detail_ws.cell(1, nc, count_header).font = openpyxl.styles.Font(bold=True)
                detail_col_map[count_header] = nc
            if findings_header not in detail_col_map:
                import openpyxl

                nc = detail_ws.max_column + 1
                detail_ws.cell(1, nc, findings_header).font = openpyxl.styles.Font(bold=True)
                detail_col_map[findings_header] = nc

            cat_findings = findings.get(cat, [])
            detail_ws.cell(detail_row, detail_col_map[count_header], len(cat_findings))
            if cat_findings:
                detail_ws.cell(
                    detail_row,
                    detail_col_map[findings_header],
                    "\n".join(f"• {f}" for f in cat_findings),
                )

        _save(wb, report_path)
    except Exception as e:
        print(f"    WARN: Could not update report for {file_stem}: {e}")
